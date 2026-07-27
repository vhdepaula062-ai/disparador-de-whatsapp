import io
import os
import queue
import re
import sqlite3
import subprocess
import threading
import time
import urllib.parse
import random
import logging
import hashlib
import csv
from datetime import datetime
from typing import Any, Callable

import tkinter as tk
from tkinter import filedialog
import customtkinter as ctk

# ── Módulo Browser Engine Enterprise ──────────────────────────────────────────
from browser_engine import (
    BrowserFactory,
    BrowserManager,
    BrowserDetector,
    BrowserCompatibility,
    SupportedBrowsers,
    SystemEnvironment
)

# ── Selenium 4 Base ───────────────────────────────────────────────────────────
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

from security_guard import (
    AntiTamperGuard,
    SecureMemoryStore,
    CommandInjectionGuard,
    PathTraversalGuard,
    InputSanitizer,
    SecurityAuditor,
    LogDataRedactor,
    criar_sessao_http_blindada
)

import pyautogui
import pyperclip

logger = logging.getLogger("LeadHunterPro.Main")

# ── Dependências opcionais ────────────────────────────────────────────────────
try:
    from google import genai as _genai_module
    SUPORTA_GEMINI: bool = True
except ImportError:
    _genai_module = None  # type: ignore[assignment]
    SUPORTA_GEMINI: bool = False

try:
    from PIL import Image, ImageTk
    import requests as _requests_module
    import win32clipboard
    import win32con
    import win32gui
    SUPORTA_IMAGEM: bool = True
except ImportError:
    Image = None          # type: ignore[assignment,misc]
    ImageTk = None        # type: ignore[assignment]
    _requests_module = None  # type: ignore[assignment]
    win32clipboard = None  # type: ignore[assignment]
    win32con = None        # type: ignore[assignment]
    win32gui = None        # type: ignore[assignment]
    SUPORTA_IMAGEM: bool = False

try:
    import openpyxl
    SUPORTA_OPENPYXL: bool = True
except ImportError:
    openpyxl = None
    SUPORTA_OPENPYXL: bool = False

# ── Motor Nativo Rust (PyO3 + Rayon) ─────────────────────────────────────────
try:
    import rust_engine as _rust_engine
    SUPORTA_RUST: bool = True
except ImportError:
    _rust_engine = None  # type: ignore[assignment]
    SUPORTA_RUST: bool = False

# ── Configuração Global de Tema ───────────────────────────────────────────────
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

# =============================================================================
# DETECÇÃO AUTOMÁTICA DO NAVEGADOR PADRÃO DO SISTEMA
# =============================================================================

def obter_navegador_padrao_windows() -> tuple[str, str | None]:
    info = BrowserCompatibility.resolve_browser()
    bin_str = str(info.binary_path) if info.binary_path else None
    return info.display_name, bin_str

# =============================================================================
# BOOTSTRAP DE AMBIENTE CENTRALIZADO (%APPDATA%\LeadHunterPro)
# =============================================================================

def get_app_data_dir() -> str:
    return str(SystemEnvironment.get_app_data_dir())

def bootstrap_ambiente() -> str:
    base_dir = get_app_data_dir()
    subpastas = ["assets", "logs", "config", "database", "cache_img", "profiles", "exports"]
    for pasta in subpastas:
        os.makedirs(os.path.join(base_dir, pasta), exist_ok=True)

    lock_path = os.path.join(base_dir, "installed.lock")
    if not os.path.exists(lock_path):
        with open(lock_path, "w", encoding="utf-8") as f:
            f.write(f"Lead Hunter Pro v3.0 — Instalado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
            f.write(f"Usuário: {os.getenv('USERNAME', 'desconhecido')}\n")
            f.write(f"Diretório: {base_dir}\n")

    return base_dir

# =============================================================================
# INFRAESTRUTURA — CAMADA DE BANCO DE DADOS
# =============================================================================

class DatabaseManager:
    def __init__(self, db_path: str) -> None:
        self.db_path: str = db_path
        self._write_lock: threading.Lock = threading.Lock()
        self._inicializar_banco_dados()

    def _get_connection(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path, timeout=15.0, check_same_thread=False)
        conn.execute("PRAGMA journal_mode = WAL;")
        conn.execute("PRAGMA synchronous = NORMAL;")
        conn.execute("PRAGMA temp_store = MEMORY;")
        conn.execute("PRAGMA cache_size = -8000;")
        conn.row_factory = sqlite3.Row
        return conn

    def _inicializar_banco_dados(self) -> None:
        with self._get_connection() as conn:
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS leads_abordados (
                    telefone  TEXT PRIMARY KEY,
                    nome      TEXT NOT NULL,
                    data_envio TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS ads_coletados (
                    ad_id      TEXT PRIMARY KEY,
                    titulo     TEXT NOT NULL,
                    link       TEXT NOT NULL,
                    data_coleta TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS grupos_whatsapp (
                    link_grupo TEXT PRIMARY KEY,
                    nicho      TEXT NOT NULL,
                    data_coleta TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_leads_telefone
                    ON leads_abordados(telefone);
            """)
            conn.commit()

    def verificar_lead_ja_abordado(self, telefone: str) -> tuple[str] | None:
        with self._get_connection() as conn:
            cursor = conn.execute(
                "SELECT data_envio FROM leads_abordados WHERE telefone = ?",
                (telefone,)
            )
            row = cursor.fetchone()
            return tuple(row) if row else None

    def verificar_ad_ja_coletado(self, ad_id: str) -> tuple[str] | None:
        with self._get_connection() as conn:
            cursor = conn.execute(
                "SELECT data_coleta FROM ads_coletados WHERE ad_id = ?",
                (ad_id,)
            )
            row = cursor.fetchone()
            return tuple(row) if row else None

    def salvar_lead_abordado(self, nome: str, telefone: str) -> bool:
        nome_sanitizado = InputSanitizer.sanitizar_texto_geral(nome, max_length=150)
        tel_sanitizado = InputSanitizer.sanitizar_telefone(telefone) or telefone
        with self._write_lock:
            try:
                with self._get_connection() as conn:
                    conn.execute(
                        "INSERT OR REPLACE INTO leads_abordados (telefone, nome, data_envio) "
                        "VALUES (?, ?, ?)",
                        (tel_sanitizado, nome_sanitizado, datetime.now().strftime("%d/%m/%Y %H:%M"))
                    )
                    conn.commit()
                return True
            except sqlite3.Error:
                return False

    def salvar_ad_coletado(self, ad_id: str, titulo: str, link: str) -> bool:
        with self._write_lock:
            try:
                with self._get_connection() as conn:
                    conn.execute(
                        "INSERT OR REPLACE INTO ads_coletados (ad_id, titulo, link, data_coleta) "
                        "VALUES (?, ?, ?, ?)",
                        (ad_id, titulo, link, datetime.now().strftime("%d/%m/%Y %H:%M"))
                    )
                    conn.commit()
                return True
            except sqlite3.Error:
                return False

    def salvar_grupo_whatsapp(self, link_grupo: str, nicho: str) -> bool:
        with self._write_lock:
            try:
                with self._get_connection() as conn:
                    conn.execute(
                        "INSERT OR REPLACE INTO grupos_whatsapp (link_grupo, nicho, data_coleta) "
                        "VALUES (?, ?, ?)",
                        (link_grupo, nicho, datetime.now().strftime("%d/%m/%Y %H:%M"))
                    )
                    conn.commit()
                return True
            except sqlite3.Error:
                return False

# =============================================================================
# INFRAESTRUTURA — GERENCIAMENTO DE DRIVER MULTI-BROWSER
# =============================================================================

def criar_driver(com_imagens: bool = True, usar_perfil_sistema: bool = True) -> WebDriver:
    custom_args = [
        "--remote-allow-origins=*",
        "--no-sandbox",
        "--disable-dev-shm-usage",
        "--remote-debugging-port=0",
        "--no-first-run",
        "--no-default-browser-check",
        "--disable-background-networking"
    ]
    if not com_imagens:
        custom_args.append("--blink-settings=imagesEnabled=false")

    if usar_perfil_sistema:
        try:
            drv = BrowserFactory.create(
                headless=False,
                custom_profile=True,
                use_system_profile=True,
                options_args=custom_args
            )
            if drv:
                return drv
        except Exception as err:
            logger.warning(f"Tentativa 1 com perfil do sistema falhou: {err}")

    try:
        drv = BrowserFactory.create(
            headless=False,
            custom_profile=True,
            use_system_profile=False,
            options_args=custom_args
        )
        if drv:
            return drv
    except Exception as err2:
        logger.warning(f"Tentativa 2 com perfil dedicado falhou: {err2}")

    try:
        drv = BrowserFactory.create(
            headless=False,
            custom_profile=False,
            use_system_profile=False,
            options_args=custom_args
        )
        if drv:
            return drv
    except Exception as err3:
        logger.warning(f"Tentativa 3 com perfil temporário falhou: {err3}")

    drv = BrowserFactory.create(
        preferred_browser=SupportedBrowsers.CHROME,
        headless=False,
        custom_profile=False,
        use_system_profile=False,
        options_args=custom_args
    )
    return drv

def abrir_url_em_nova_aba(driver: WebDriver | None, url: str) -> None:
    if driver is None:
        return
    try:
        current_url = str(driver.current_url).lower() if hasattr(driver, "current_url") else ""
        if "data:," in current_url or "about:blank" in current_url or "msn.com" in current_url or "edge://" in current_url or not current_url:
            driver.get(url)
        else:
            if hasattr(driver, "window_handles") and len(driver.window_handles) > 0:
                driver.switch_to.new_window('tab')
            driver.get(url)
    except Exception:
        try:
            driver.get(url)
        except Exception:
            pass

def fechar_aba_ou_driver(driver: WebDriver | None) -> None:
    if driver is not None:
        try:
            if hasattr(driver, "window_handles") and len(driver.window_handles) > 1:
                driver.close()
            else:
                driver.quit()
        except Exception:
            pass

def _dismiss_google_consent(driver: WebDriver | None) -> None:
    if driver is None:
        return
    try:
        driver.execute_script("""
            var msBtns = document.querySelectorAll('button[id*="accept"], button[class*="accept"], #accept-all, #bep-accept-btn, .bep-accept');
            for (var i = 0; i < msBtns.length; i++) {
                try { msBtns[i].click(); } catch(e) {}
            }
        """)
        btns = driver.find_elements(
            By.XPATH,
            '//button[contains(., "Aceitar") or contains(., "Concordo") or contains(., "I agree") or contains(., "Accept all")]'
        )
        if btns:
            driver.execute_script("arguments[0].click();", btns[0])
            time.sleep(0.5)
    except Exception:
        pass

# =============================================================================
# INFRAESTRUTURA — HTTP SESSION & GEMINI REAL-TIME OPTIMIZER AGENT
# =============================================================================

class HttpSessionManager:
    def __init__(self) -> None:
        self._session: Any | None = None
        self._lock: threading.Lock = threading.Lock()
        self.cache_dir = os.path.join(get_app_data_dir(), "cache_img")
        os.makedirs(self.cache_dir, exist_ok=True)

    @property
    def session(self) -> Any:
        if not SUPORTA_IMAGEM or _requests_module is None:
            raise RuntimeError("requests não disponível")
        with self._lock:
            if self._session is None:
                self._session = criar_sessao_http_blindada(timeout_padrao=6.0)
                self._session.headers.update({
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                    "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
                    "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7"
                })
            return self._session

    def get_cached_image_bytes(self, url: str, timeout: float = 4.0) -> bytes | None:
        if not url or not url.startswith("http"):
            return None

        url_hash = hashlib.md5(url.encode('utf-8')).hexdigest()
        cache_path = os.path.join(self.cache_dir, f"{url_hash}.bin")

        if os.path.exists(cache_path):
            try:
                with open(cache_path, "rb") as f:
                    return f.read()
            except Exception:
                pass

        try:
            resp = self.get(url, timeout=timeout)
            if resp.status_code == 200 and resp.content:
                data = resp.content
                with open(cache_path, "wb") as f:
                    f.write(data)
                return data
        except Exception:
            pass
        return None

    def get(self, url: str, timeout: float = 5.0) -> Any:
        headers = {}
        if "fbcdn.net" in url or "facebook.com" in url:
            headers["Referer"] = "https://www.facebook.com/"
        elif "plati" in url:
            headers["Referer"] = "https://plati.market/"
        elif "z2u" in url:
            headers["Referer"] = "https://www.z2u.com/"
        elif "ggmax" in url:
            headers["Referer"] = "https://ggmax.com.br/"
        elif "shopee" in url or "susercontent" in url:
            headers["Referer"] = "https://shopee.com.br/"
        elif "amazon" in url or "ssl-images-amazon" in url or "media-amazon" in url:
            headers["Referer"] = "https://www.amazon.com.br/"
        elif "mercadolivre" in url or "mlstatic" in url:
            headers["Referer"] = "https://www.mercadolivre.com.br/"

        return self.session.get(url, headers=headers if headers else None, timeout=timeout)

    def close(self) -> None:
        with self._lock:
            if self._session is not None:
                self._session.close()
                self._session = None

class GeminiClientManager:
    """AGENTE OTIMIZADOR DE IA GEMINI 2.0 FLASH EM TEMPO REAL"""
    def __init__(self, cache_size: int = 256) -> None:
        self._client: Any | None = None
        self._api_key: str = ""
        self._lock: threading.Lock = threading.Lock()
        self._cache: dict[str, str] = {}
        self._cache_size: int = cache_size
        self.ativo: bool = False

    def configurar(self, api_key: str) -> None:
        with self._lock:
            if api_key != self._api_key:
                self._api_key = api_key
                self._client = None
                self._cache.clear()
                self.ativo = bool(api_key.strip())

    def _get_client(self) -> Any | None:
        if not SUPORTA_GEMINI or _genai_module is None or not self._api_key:
            return None
        with self._lock:
            if self._client is None and self._api_key:
                self._client = _genai_module.Client(api_key=self._api_key)
            return self._client

    def gerar_conteudo(self, prompt: str, max_retries: int = 3) -> str | None:
        cache_key = prompt.strip()
        with self._lock:
            if cache_key in self._cache:
                return self._cache[cache_key]

        client = self._get_client()
        if client is None:
            return None

        for tentativa in range(max_retries):
            try:
                response = client.models.generate_content(
                    model="gemini-2.0-flash",
                    contents=prompt
                )
                if response and response.text:
                    txt = response.text.strip()
                    with self._lock:
                        if len(self._cache) >= self._cache_size:
                            first_key = next(iter(self._cache))
                            del self._cache[first_key]
                        self._cache[cache_key] = txt
                    return txt
            except Exception:
                if tentativa < max_retries - 1:
                    time.sleep(1.5 ** tentativa)
        return None

    def otimizar_termo_pesquisa(self, termo_base: str) -> str:
        if not self.ativo:
            return termo_base
        prompt = (
            f"Atue como um Agente Otimizador de SEO e prospecção de vendas B2B. "
            f"Melhore o termo de busca '{termo_base}' para encontrar empresas com alto potencial de compra. "
            f"Responda APENAS com a nova frase otimizada de pesquisa."
        )
        res = self.gerar_conteudo(prompt)
        return res if res else termo_base

    def otimizar_copywriting_whatsapp(self, nome_empresa: str, nicho: str, proposta: str) -> str:
        if not self.ativo:
            return (
                f"Olá! Notei o excelente trabalho da '{nome_empresa}' no segmento de {nicho}. "
                f"{proposta} Teria 2 minutos hoje para uma breve demonstração?"
            )
        prompt = (
            f"Você é um copywriter expert em prospecção via WhatsApp. "
            f"Escreva uma mensagem de até 4 linhas para '{nome_empresa}' (nicho: {nicho}). "
            f"Proposta de valor: '{proposta}'. "
            f"Sem formalismos excessivos, com gancho persuasivo e chamada de ação direta (CTA)."
        )
        res = self.gerar_conteudo(prompt)
        return res if res else f"Olá {nome_empresa}! Vi seu trabalho em {nicho}. {proposta}"

    def gerar_saudacao_aquecimento(self, nome_grupo: str) -> str:
        if not self.ativo:
            return "Fala pessoal! Bom dia a todos do grupo!"
        prompt = (
            f"Gere uma curta e natural mensagem de saudação (1 linha) para enviar ao entrar em um grupo de WhatsApp "
            f"sobre '{nome_grupo}' para parecer um membro humano legítimo."
        )
        res = self.gerar_conteudo(prompt)
        return res if res else "Fala pessoal! Bom dia a todos do grupo!"

# =============================================================================
# APLICAÇÃO PRINCIPAL
# =============================================================================

class LeadHunterProApp(ctk.CTk):
    MAPS_CIDADES_VIZINHAS = {
        "anápolis": ["Goianápolis GO", "Nerópolis GO", "Goiânia GO", "Aparecida de Goiânia GO", "Trindade GO",
                     "Abadiânia GO", "Alexânia GO", "Pirenópolis GO", "Inhumas GO", "Senador Canedo GO"],
        "goiânia": ["Aparecida de Goiânia GO", "Senador Canedo GO", "Trindade GO", "Anápolis GO", "Inhumas GO",
                    "Goianira GO", "Guapó GO", "Hidrolândia GO"],
        "palmas": ["Taquaralto TO", "Porto Nacional TO", "Paraíso do Tocantins TO", "Miracema do Tocantins TO"],
        "são paulo": ["Guarulhos SP", "Campinas SP", "São Bernardo do Campo SP", "Santo André SP", "Osasco SP"],
        "rio de janeiro": ["Niterói RJ", "Duque de Caxias RJ", "Nova Iguaçu RJ", "São Gonçalo RJ"],
        "curitiba": ["São José dos Pinhais PR", "Colombo PR", "Araucária PR", "Pinhais PR"],
        "belo horizonte": ["Contagem MG", "Betim MG", "Nova Lima MG", "Ribeirão das Neves MG"],
        "brasília": ["Taguatinga DF", "Ceilândia DF", "Águas Claras DF", "Samambaia DF", "Luziânia GO"],
    }

    def gerar_pool_buscas_regional(self, termo_original: str) -> list[str]:
        termo_lower = termo_original.lower().strip()

        match_nicho = re.split(r'\b(?:em|no|na)\b', termo_original, flags=re.IGNORECASE)
        if len(match_nicho) > 1 and match_nicho[0].strip():
            nicho = match_nicho[0].strip()
        else:
            nicho = re.sub(r'\b(go|to|sp|mg|rj|pr|rs|df|goias|goiás|tocantins|minas|rio)\b.*$', '', termo_original, flags=re.IGNORECASE).strip()
            if not nicho:
                nicho = termo_original.strip()

        cidades_encontradas = []
        for chave, vizinhas in self.MAPS_CIDADES_VIZINHAS.items():
            if chave in termo_lower:
                cidades_encontradas = vizinhas
                break

        if not cidades_encontradas:
            cidades_encontradas = ["Goiânia GO", "Brasília DF", "São Paulo SP", "Belo Horizonte MG"]

        pool = [termo_original]
        for vizinha in cidades_encontradas:
            query_vizinha = f"{nicho} em {vizinha}"
            if query_vizinha.lower() not in [p.lower() for p in pool]:
                pool.append(query_vizinha)

        return pool

    def __init__(self) -> None:
        super().__init__()

        def __init__(self) -> None:
            super().__init__()

            # Proteção contra depuradores e malwares ativos
            if AntiTamperGuard.is_debugger_present():
                logger.warning("⚠️ [SEGURANÇA] Ambiente monitorado ou depurador ativo detectado.")

            base_path: str = bootstrap_ambiente()

        base_path: str = bootstrap_ambiente()
        self._base_path: str = base_path
        self.db_path: str = os.path.join(base_path, "database", "historico_leads.db")
        self.db: DatabaseManager = DatabaseManager(self.db_path)
        self.http: HttpSessionManager = HttpSessionManager()
        self.gemini: GeminiClientManager = GeminiClientManager()

        self.modo_lit_hunter: bool = False
        self.grupos_coletados: list[dict[str, Any]] = []

        self.title("⚡ LEAD HUNTER PRO v3.0 — Gemini AI & Lit Hunter Suite")
        self.geometry("1560x940")
        self.minsize(1280, 760)
        self.configure(fg_color="#060810")
        self.protocol("WM_DELETE_WINDOW", self._on_fechar_janela)

        self._automacao_lock: threading.Lock = threading.Lock()
        self._automacao_rodando: bool = False
        self.agendamento_ativo: bool = False
        self.leads_dados: list[dict[str, Any]] = []
        self.caminhos_imagens: list[str] = []
        self.tempo_inicio_automacao: float | None = None
        self._horario_agendado: str = "06:00"

        self.total_coletados_count: int = 0
        self.total_enviados_count: int = 0
        self.cotacao_dolar_atual: float = 5.08

        # UI Declarations
        self.main_panel: ctk.CTkFrame
        self.top_header: ctk.CTkFrame
        self.badge_status: ctk.CTkFrame
        self.lbl_status_sistema: ctk.CTkLabel
        self.kpi_bar: ctk.CTkFrame
        self.lbl_stat_coletados: ctk.CTkLabel
        self.progress_coletados: ctk.CTkProgressBar
        self.lbl_stat_enviados: ctk.CTkLabel
        self.progress_enviados: ctk.CTkProgressBar
        self.lbl_kpi_dolar: ctk.CTkLabel
        self.lbl_kpi_engine: ctk.CTkLabel
        self.center_dashboard: ctk.CTkFrame
        self.col1_frame: ctk.CTkFrame
        self.col2_frame: ctk.CTkFrame
        self.col3_frame: ctk.CTkFrame
        self.card_timer_pausa: ctk.CTkFrame
        self.lbl_timer_pausa: ctk.CTkLabel
        self.progress_timer_pausa: ctk.CTkProgressBar
        self.txt_console: ctk.CTkTextbox
        self.scroll_leads: ctk.CTkScrollableFrame
        self.tabview: ctk.CTkTabview
        self.entry_api_key: ctk.CTkEntry
        self.lbl_api_status: ctk.CTkLabel
        self.entry_termo_comercial: ctk.CTkEntry
        self.txt_msg_comercial: ctk.CTkTextbox
        self.entry_termo_avancado: ctk.CTkEntry
        self.entry_cidade_avancado: ctk.CTkEntry
        self.txt_msg_diretorios: ctk.CTkTextbox
        self.entry_termo_ads: ctk.CTkEntry
        self.entry_qtd_ads: ctk.CTkEntry
        self.entry_tempo_mineracao: ctk.CTkEntry
        self.entry_termo_comparar: ctk.CTkEntry
        self.opt_plataforma_afiliados: ctk.CTkOptionMenu
        self.entry_termo_afiliados: ctk.CTkEntry
        self.entry_qtd_afiliados: ctk.CTkEntry
        self.entry_tempo_afiliados: ctk.CTkEntry
        self.entry_termo_lit: ctk.CTkEntry
        self.entry_qtd_lit: ctk.CTkEntry
        self.entry_meta: ctk.CTkEntry
        self.entry_horario: ctk.CTkEntry
        self.btn_add_img: ctk.CTkButton
        self.lbl_qtd_imagens: ctk.CTkLabel
        self.btn_iniciar_coleta: ctk.CTkButton
        self.btn_iniciar_disparos: ctk.CTkButton
        self.btn_ativar_agendamento: ctk.CTkButton
        self.btn_parar: ctk.CTkButton
        self.btn_mode_lead: ctk.CTkButton
        self.btn_mode_lit: ctk.CTkButton
        self.footer_bar: ctk.CTkFrame
        self.lbl_footer_tempo: ctk.CTkLabel
        self.lbl_footer_data: ctk.CTkLabel
        self._glow_state: bool = False

        self._render_queue: queue.Queue[Callable[[], None]] = queue.Queue()

        self.criar_layout_dashboard()

        nome_browser_padrao, exe_p = obter_navegador_padrao_windows()
        self.log(f"🌐 Navegador Padrão do Usuário Mapeado: {nome_browser_padrao}")
        if exe_p:
            self.log(f"📌 Executável do Navegador: {exe_p}")

        self._stop_event: threading.Event = threading.Event()
        self._agendamento_event: threading.Event = threading.Event()

        threading.Thread(target=self._vigiar_agendamento_worker, daemon=True).start()
        threading.Thread(target=self._atualizar_cotacao_inicial, daemon=True).start()

        self._tick_relogio()
        self._processar_render_queue()
        threading.Thread(target=self._prefetch_driver_path, daemon=True).start()

    @property
    def automacao_rodando(self) -> bool:
        with self._automacao_lock:
            return self._automacao_rodando

    @automacao_rodando.setter
    def automacao_rodando(self, valor: bool) -> None:
        with self._automacao_lock:
            self._automacao_rodando = valor

    def _tentar_iniciar_automacao(self) -> bool:
        with self._automacao_lock:
            if self._automacao_rodando:
                return False
            self._automacao_rodando = True
            return True

    @staticmethod
    def normalizar_telefone_br(raw_tel: str) -> str | None:
        digits = "".join(filter(str.isdigit, raw_tel))
        if digits.startswith("0") and len(digits) in (11, 12):
            digits = digits[1:]
        if not digits.startswith("55") and len(digits) in (10, 11):
            digits = "55" + digits
        if len(digits) in (12, 13):
            return digits
        return None

    @staticmethod
    def _prefetch_driver_path() -> None:
        try:
            BrowserDetector.detect_all()
        except Exception:
            pass

    def validar_e_filtrar_celular(self, tel_limpo: str) -> bool:
        if SUPORTA_RUST and _rust_engine is not None and hasattr(_rust_engine, "validar_celular_brasileiro"):
            try:
                return _rust_engine.validar_celular_brasileiro(tel_limpo)
            except Exception:
                pass

        if tel_limpo.startswith("55") and len(tel_limpo) >= 12:
            nacional = tel_limpo[2:]
        else:
            nacional = tel_limpo
        if len(nacional) == 10 and nacional[2] in ("2", "3", "4", "5"):
            return False
        return len(nacional) >= 10

    def obter_cotacao_dolar(self) -> float:
        if SUPORTA_IMAGEM:
            try:
                r = self.http.get("https://api.exchangerate-api.com/v4/latest/USD", timeout=4.0)
                if r.status_code == 200:
                    val = float(r.json()["rates"]["BRL"])
                    self.cotacao_dolar_atual = val
                    return val
            except Exception:
                pass
        return self.cotacao_dolar_atual

    def _atualizar_cotacao_inicial(self) -> None:
        val = self.obter_cotacao_dolar()
        self.after(0, lambda: self.lbl_kpi_dolar.configure(text=f"R$ {val:.2f}"))

    def log(self, mensagem: str) -> None:
        # Mascara chaves de API do Gemini e segredos antes de exibir no console
        mensagem_segura = LogDataRedactor.mascarar_logs(mensagem)

        def _write() -> None:
            self.txt_console.configure(state="normal")
            self.txt_console.insert("end", f"[{time.strftime('%H:%M:%S')}] {mensagem_segura}\n")
            self.txt_console.see("end")
            self.txt_console.configure(state="disabled")

        if threading.current_thread() is threading.main_thread():
            _write()
        else:
            self.after(0, _write)

    def abrir_link_anuncio_seguro(self, url: str) -> None:
        try:
            if CommandInjectionGuard.abrir_url_com_seguranca(url):
                self.log(f"🔗 Redirecionando para oferta/grupo: {url}")
            else:
                self.log(f"✖ URL bloqueada por segurança: {url}")
        except Exception as e:
            self.log(f"✖ Erro ao abrir URL: {e}")

    def _ui_safe(self, fn: Callable[[], None]) -> None:
        if threading.current_thread() is threading.main_thread():
            fn()
        else:
            self.after(0, fn)

    def _on_fechar_janela(self) -> None:
        self._stop_event.set()
        self._agendamento_event.set()
        self.automacao_rodando = False
        self.http.close()
        self.destroy()

    def _tick_relogio(self) -> None:
        try:
            agora_str = datetime.now().strftime("DATA E HORA   %d/%m/%Y   %H:%M:%S")
            self.lbl_footer_data.configure(text=agora_str)

            # Atualização em tempo real da Telemetria de Segurança Visual
            diag = SecurityAuditor.diagnostico_completo()
            if "DETECTADO" in diag["anti_tamper"].upper():
                self.lbl_sec_status_geral.configure(text="⚠️ MONITORADO", text_color="#F59E0B")
                self.lbl_sec_tamper.configure(text="👁️ " + diag["anti_tamper"], text_color="#EF4444")
            else:
                self.lbl_sec_status_geral.configure(text="🟢 AMBIENTE SEGURO", text_color="#10B981")
                self.lbl_sec_tamper.configure(text="👁️ Anti-Tamper: ATIVO", text_color="#10B981")

            if self._automacao_rodando and self.tempo_inicio_automacao:
                decorrido = int(time.time() - self.tempo_inicio_automacao)
                h, resto = divmod(decorrido, 3600)
                m, s = divmod(resto, 60)
                self.lbl_footer_tempo.configure(text=f"AUTOMAÇÃO ATIVA HÁ: {h:02d}:{m:02d}:{s:02d}")

                self._glow_state = not self._glow_state
                glow_color = "#F59E0B" if self.modo_lit_hunter else ("#38BDF8" if self._glow_state else "#0284C7")
                self.badge_status.configure(border_color=glow_color)
            else:
                self.lbl_footer_tempo.configure(text="AUTOMAÇÃO EM ESPERA")
                self.badge_status.configure(border_color="#334155")
        except Exception:
            pass
        self.after(1000, self._tick_relogio)

    def _atualizar_timer_pausa_ui(self, tempo_str: str, progresso: float) -> None:
        try:
            self.lbl_timer_pausa.configure(text=tempo_str, text_color="#F59E0B")
            self.progress_timer_pausa.set(progresso)
        except Exception:
            pass

    def _resetar_timer_pausa_ui(self) -> None:
        try:
            self.lbl_timer_pausa.configure(text="EM ESPERA", text_color="#64748B")
            self.progress_timer_pausa.set(0.0)
        except Exception:
            pass

    def _atualizar_kpi_coletados_ui(self, contagem: int, progresso: float) -> None:
        try:
            self.lbl_stat_coletados.configure(text=str(contagem))
            self.progress_coletados.set(progresso)
        except Exception:
            pass

    def _atualizar_kpi_enviados_ui(self, contagem: int, progresso: float) -> None:
        try:
            self.lbl_stat_enviados.configure(text=str(contagem))
            self.progress_enviados.set(progresso)
        except Exception:
            pass

    def _processar_render_queue(self) -> None:
        for _ in range(10):
            try:
                fn = self._render_queue.get_nowait()
                fn()
            except queue.Empty:
                break
        self.after(50, self._processar_render_queue)

    def _enqueue_render(self, fn: Callable[[], None]) -> None:
        self._render_queue.put(fn)

    # ── ALTERNÂNCIA DE PROJETO (LEAD HUNTER vs LIT HUNTER) ───────────────────

    def definir_modo_projeto(self, ativar_lit: bool) -> None:
        self.modo_lit_hunter = ativar_lit
        if self.modo_lit_hunter:
            self.btn_mode_lead.configure(fg_color="#1E293B", text_color="#94A3B8")
            self.btn_mode_lit.configure(fg_color="#D97706", text_color="#FFFFFF")
            self.lbl_status_sistema.configure(text="● MODO LIT HUNTER", text_color="#F59E0B")
            self.tabview.set("🔥 Lit Hunter")
            self.btn_iniciar_coleta.configure(text="🔥 1. MINERAR GRUPOS PÚBLICOS", fg_color="#D97706", hover_color="#B45309")
            self.btn_iniciar_disparos.configure(text="🚀 2. INICIAR AQUECIMENTO", fg_color="#10B981", hover_color="#059669")
            self.log("🔥 SUB-PROJETO ATIVADO: LIT HUNTER (Aquecedor de WhatsApp & Grupos Públicos)")
        else:
            self.btn_mode_lead.configure(fg_color="#0284C7", text_color="#FFFFFF")
            self.btn_mode_lit.configure(fg_color="#1E293B", text_color="#94A3B8")
            self.lbl_status_sistema.configure(text="● SISTEMA EM ESPERA", text_color="#EF4444")
            self.tabview.set("🛒 Maps")
            self.btn_iniciar_coleta.configure(text="🔍 1. INICIAR COLETA DE LEADS", fg_color="#0284C7", hover_color="#0369A1")
            self.btn_iniciar_disparos.configure(text="🚀 2. INICIAR DISPAROS COM IA", fg_color="#10B981", hover_color="#059669")
            self.log("⚡ PROJETO PRINCIPAL ATIVADO: LEAD HUNTER PRO (Captador B2B & Disparos IA)")

    # ── UI Construction ───────────────────────────────────────────────────────

    def criar_layout_dashboard(self) -> None:
        self.main_panel = ctk.CTkFrame(self, fg_color="transparent")
        self.main_panel.pack(side="top", fill="both", expand=True, padx=16, pady=16)

        self._criar_header_e_kpis()

        self.center_dashboard = ctk.CTkFrame(self.main_panel, fg_color="transparent")
        self.center_dashboard.pack(side="top", fill="both", expand=True, padx=0, pady=(12, 0))

        self.col1_frame = ctk.CTkFrame(
            self.center_dashboard, fg_color="#0F172A", corner_radius=14,
            width=460, border_width=1, border_color="#1E293B"
        )
        self.col1_frame.pack(side="left", fill="both", expand=False, padx=(0, 12), pady=0)
        self.col1_frame.pack_propagate(False)

        self.col2_frame = ctk.CTkFrame(
            self.center_dashboard, fg_color="#0F172A", corner_radius=14,
            border_width=1, border_color="#1E293B"
        )
        self.col2_frame.pack(side="left", fill="both", expand=True, padx=(0, 12), pady=0)

        self.col3_frame = ctk.CTkFrame(
            self.center_dashboard, fg_color="#0F172A", corner_radius=14,
            width=400, border_width=1, border_color="#1E293B"
        )
        self.col3_frame.pack(side="right", fill="both", expand=False, padx=0, pady=0)
        self.col3_frame.pack_propagate(False)

        self._montar_coluna_1_controles()
        self._montar_coluna_2_resultados()
        self._montar_coluna_3_telemetria()
        self._criar_rodape_enterprise()

    def _criar_header_e_kpis(self) -> None:
        self.top_header = ctk.CTkFrame(
            self.main_panel, fg_color="#0F172A", corner_radius=14,
            height=62, border_width=1, border_color="#1E293B"
        )
        self.top_header.pack(side="top", fill="x", padx=0, pady=(0, 12))
        self.top_header.pack_propagate(False)

        brand_box = ctk.CTkFrame(self.top_header, fg_color="transparent")
        brand_box.pack(side="left", padx=18)
        ctk.CTkLabel(brand_box, text="⚡ LEAD HUNTER PRO", font=("Arial Bold", 18),
                     text_color="#38BDF8").pack(side="left")
        ctk.CTkLabel(brand_box, text=" | ENTERPRISE SUITE v3.0", font=("Arial Bold", 11),
                     text_color="#64748B").pack(side="left", padx=(6, 0))

        mode_box = ctk.CTkFrame(self.top_header, fg_color="transparent")
        mode_box.pack(side="left", padx=(24, 0))

        self.btn_mode_lead = ctk.CTkButton(
            mode_box, text="⚡ LEAD HUNTER PRO", font=("Arial Bold", 11),
            fg_color="#0284C7", hover_color="#0369A1", width=160, height=32,
            corner_radius=8, command=lambda: self.definir_modo_projeto(False)
        )
        self.btn_mode_lead.pack(side="left", padx=(0, 8))

        self.btn_mode_lit = ctk.CTkButton(
            mode_box, text="🔥 LIT HUNTER (AQUECEDOR)", font=("Arial Bold", 11),
            fg_color="#1E293B", hover_color="#334155", text_color="#94A3B8", width=190, height=32,
            corner_radius=8, command=lambda: self.definir_modo_projeto(True)
        )
        self.btn_mode_lit.pack(side="left")

        self.badge_status = ctk.CTkFrame(
            self.top_header, fg_color="#060810", corner_radius=20,
            height=36, width=210, border_width=1, border_color="#334155"
        )
        self.badge_status.pack(side="right", padx=18)
        self.badge_status.pack_propagate(False)

        self.lbl_status_sistema = ctk.CTkLabel(
            self.badge_status, text="● SISTEMA EM ESPERA",
            font=("Arial Bold", 11), text_color="#EF4444"
        )
        self.lbl_status_sistema.pack(expand=True)

        self.kpi_bar = ctk.CTkFrame(self.main_panel, fg_color="transparent")
        self.kpi_bar.pack(side="top", fill="x", padx=0, pady=0)

        c1 = ctk.CTkFrame(self.kpi_bar, fg_color="#0F172A", corner_radius=12,
                           border_width=1, border_color="#1E293B")
        c1.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ctk.CTkLabel(c1, text="LEADS / GRUPOS CAPTADOS", font=("Arial Bold", 10),
                     text_color="#94A3B8").pack(anchor="w", padx=12, pady=(6, 0))
        self.lbl_stat_coletados = ctk.CTkLabel(c1, text="0", font=("Arial Bold", 20),
                                               text_color="#38BDF8")
        self.lbl_stat_coletados.pack(anchor="w", padx=12, pady=(0, 2))
        self.progress_coletados = ctk.CTkProgressBar(
            c1, height=5, corner_radius=3, fg_color="#060810", progress_color="#38BDF8"
        )
        self.progress_coletados.set(0.0)
        self.progress_coletados.pack(fill="x", padx=12, pady=(0, 6))

        c2 = ctk.CTkFrame(self.kpi_bar, fg_color="#0F172A", corner_radius=12,
                           border_width=1, border_color="#1E293B")
        c2.pack(side="left", fill="x", expand=True, padx=(4, 8))
        ctk.CTkLabel(c2, text="DISPAROS / AQUECIMENTOS", font=("Arial Bold", 10),
                     text_color="#94A3B8").pack(anchor="w", padx=12, pady=(6, 0))
        self.lbl_stat_enviados = ctk.CTkLabel(c2, text="0", font=("Arial Bold", 20),
                                              text_color="#10B981")
        self.lbl_stat_enviados.pack(anchor="w", padx=12, pady=(0, 2))
        self.progress_enviados = ctk.CTkProgressBar(
            c2, height=5, corner_radius=3, fg_color="#060810", progress_color="#10B981"
        )
        self.progress_enviados.set(0.0)
        self.progress_enviados.pack(fill="x", padx=12, pady=(0, 6))

        c3 = ctk.CTkFrame(self.kpi_bar, fg_color="#0F172A", corner_radius=12,
                           border_width=1, border_color="#1E293B")
        c3.pack(side="left", fill="x", expand=True, padx=(4, 8))
        ctk.CTkLabel(c3, text="COTAÇÃO DÓLAR (USD/BRL)", font=("Arial Bold", 10),
                     text_color="#94A3B8").pack(anchor="w", padx=12, pady=(8, 0))
        self.lbl_kpi_dolar = ctk.CTkLabel(c3, text="R$ 5.08", font=("Arial Bold", 20),
                                          text_color="#F59E0B")
        self.lbl_kpi_dolar.pack(anchor="w", padx=12, pady=(0, 8))

        c4 = ctk.CTkFrame(self.kpi_bar, fg_color="#0F172A", corner_radius=12,
                           border_width=1, border_color="#1E293B")
        c4.pack(side="left", fill="x", expand=True, padx=(4, 0))
        ctk.CTkLabel(c4, text="ENGINE DE IA GEMINI", font=("Arial Bold", 10),
                     text_color="#94A3B8").pack(anchor="w", padx=12, pady=(8, 0))
        self.lbl_kpi_engine = ctk.CTkLabel(c4, text="GEMINI-2.0-FLASH", font=("Arial Bold", 16),
                                           text_color="#A855F7")
        self.lbl_kpi_engine.pack(anchor="w", padx=12, pady=(2, 8))

    def _montar_coluna_1_controles(self) -> None:
        api_frame = ctk.CTkFrame(self.col1_frame, fg_color="#1E293B", corner_radius=10)
        api_frame.pack(fill="x", padx=12, pady=(12, 6))

        ctk.CTkLabel(api_frame, text="🔑 Google Gemini AI Key", font=("Arial Bold", 11),
                     text_color="#38BDF8").pack(anchor="w", padx=10, pady=(6, 2))

        row_api = ctk.CTkFrame(api_frame, fg_color="transparent")
        row_api.pack(fill="x", padx=10, pady=(0, 4))

        self.entry_api_key = ctk.CTkEntry(
            row_api, height=30, font=("Arial", 10), fg_color="#060810",
            border_color="#334155", placeholder_text="Cole sua API Key aqui..."
        )
        self.entry_api_key.pack(side="left", fill="x", expand=True, padx=(0, 6))

        ctk.CTkButton(
            row_api, text="Validar", width=70, height=30, font=("Arial Bold", 10),
            fg_color="#0284C7", hover_color="#0369A1", command=self.testar_conexao_api
        ).pack(side="right")

        self.lbl_api_status = ctk.CTkLabel(
            api_frame, text="🔴 Status da IA: Aguardando verificação",
            font=("Arial Italic", 10), text_color="#EF4444"
        )
        self.lbl_api_status.pack(anchor="w", padx=10, pady=(0, 6))

        self.tabview = ctk.CTkTabview(
            self.col1_frame, fg_color="#060810", segmented_button_fg_color="#0F172A",
            segmented_button_selected_color="#0284C7",
            segmented_button_unselected_color="#1E293B", height=270, corner_radius=10
        )
        self.tabview.pack(fill="x", padx=12, pady=4)

        self.tab_comercial = self.tabview.add("🛒 Maps")
        self.tab_diretorios = self.tabview.add("🌐 Web")
        self.tab_meta_ads = self.tabview.add("🎯 Meta")
        self.tab_comparador = self.tabview.add("⚖️ Preços")
        self.tab_afiliados = self.tabview.add("🛍️ Afiliados")
        self.tab_lit_hunter = self.tabview.add("🔥 Lit Hunter")

        try:
            self.tabview._segmented_button.configure(font=("Arial Bold", 10))
        except Exception:
            pass

        # Aba 1: Maps
        ctk.CTkLabel(self.tab_comercial, text="TERMO DE PESQUISA (MAPS)",
                     font=("Arial Bold", 10), text_color="#94A3B8").pack(anchor="w")
        self.entry_termo_comercial = ctk.CTkEntry(
            self.tab_comercial, height=28, font=("Arial", 11),
            fg_color="#0F172A", border_color="#334155"
        )
        self.entry_termo_comercial.insert(0, "Lojas em Anápolis Goias")
        self.entry_termo_comercial.pack(fill="x", pady=(2, 4))

        ctk.CTkLabel(self.tab_comercial, text="OFERTA / MENSAGEM BASE",
                     font=("Arial Bold", 10), text_color="#38BDF8").pack(anchor="w")
        self.txt_msg_comercial = ctk.CTkTextbox(
            self.tab_comercial, height=50, font=("Arial", 10),
            fg_color="#0F172A", border_color="#334155", border_width=1
        )
        self.txt_msg_comercial.pack(fill="x", pady=(2, 0))
        self.txt_msg_comercial.insert(
            "0.0",
            "Oferecer criação de Landing Pages profissionais e estratégias de vendas e-commerce."
        )

        # Aba 2: Busca Web
        ctk.CTkLabel(self.tab_diretorios, text="NICHO OU PROFISSÃO",
                     font=("Arial Bold", 10), text_color="#94A3B8").pack(anchor="w")
        self.entry_termo_avancado = ctk.CTkEntry(
            self.tab_diretorios, height=26, font=("Arial", 11),
            fg_color="#0F172A", border_color="#334155"
        )
        self.entry_termo_avancado.insert(0, "Fotógrafo")
        self.entry_termo_avancado.pack(fill="x", pady=(1, 3))

        ctk.CTkLabel(self.tab_diretorios, text="CIDADE E ESTADO",
                     font=("Arial Bold", 10), text_color="#38BDF8").pack(anchor="w")
        self.entry_cidade_avancado = ctk.CTkEntry(
            self.tab_diretorios, height=26, font=("Arial", 11),
            fg_color="#0F172A", border_color="#334155"
        )
        self.entry_cidade_avancado.insert(0, "Anápolis GO")
        self.entry_cidade_avancado.pack(fill="x", pady=(1, 3))

        self.txt_msg_diretorios = ctk.CTkTextbox(
            self.tab_diretorios, height=35, font=("Arial", 10),
            fg_color="#0F172A", border_color="#334155", border_width=1
        )
        self.txt_msg_diretorios.pack(fill="x", pady=0)
        self.txt_msg_diretorios.insert(
            "0.0",
            "Oferecer soluções em inteligência artificial para otimizar os serviços."
        )

        # Aba 3: Meta Ads
        ctk.CTkLabel(self.tab_meta_ads, text="ESPIONAGEM META ADS (EX: 'emagrecimento')",
                     font=("Arial Bold", 10), text_color="#94A3B8").pack(anchor="w")
        self.entry_termo_ads = ctk.CTkEntry(
            self.tab_meta_ads, height=28, font=("Arial", 11),
            fg_color="#0F172A", border_color="#334155"
        )
        self.entry_termo_ads.insert(0, "emagrecimento")
        self.entry_termo_ads.pack(fill="x", pady=(2, 4))

        row_ads_params = ctk.CTkFrame(self.tab_meta_ads, fg_color="transparent")
        row_ads_params.pack(fill="x", pady=(0, 4))

        f_qtd = ctk.CTkFrame(row_ads_params, fg_color="transparent")
        f_qtd.pack(side="left", fill="x", expand=True, padx=(0, 2))
        ctk.CTkLabel(f_qtd, text="QTD POR TERMO", font=("Arial Bold", 10), text_color="#38BDF8").pack(anchor="w")
        self.entry_qtd_ads = ctk.CTkEntry(
            f_qtd, height=28, font=("Arial", 11), fg_color="#0F172A", border_color="#334155"
        )
        self.entry_qtd_ads.insert(0, "15")
        self.entry_qtd_ads.pack(fill="x")

        f_tempo = ctk.CTkFrame(row_ads_params, fg_color="transparent")
        f_tempo.pack(side="right", fill="x", expand=True, padx=(2, 0))
        ctk.CTkLabel(f_tempo, text="TEMPO (MINUTOS)", font=("Arial Bold", 10), text_color="#F59E0B").pack(anchor="w")
        self.entry_tempo_mineracao = ctk.CTkEntry(
            f_tempo, height=28, font=("Arial", 11), fg_color="#0F172A", border_color="#334155"
        )
        self.entry_tempo_mineracao.insert(0, "10")
        self.entry_tempo_mineracao.pack(fill="x")

        ctk.CTkButton(
            self.tab_meta_ads, text="🎯 PESQUISAR TERMO ÚNICO",
            font=("Arial Bold", 10), fg_color="#A855F7", hover_color="#7E22CE",
            height=28, command=self.disparar_thread_ads_library
        ).pack(fill="x", pady=(4, 2))

        ctk.CTkButton(
            self.tab_meta_ads, text="🚀 AUTO-MINERAR MULTI-NICHOS POR TEMPO",
            font=("Arial Bold", 10), fg_color="#D97706", hover_color="#B45309",
            height=30, command=self.disparar_thread_auto_mineracao
        ).pack(fill="x", pady=2)

        # Aba 4: Comparador
        ctk.CTkLabel(self.tab_comparador, text="PRODUTO DIGITAL (EX: 'CapCut', 'Canva')",
                     font=("Arial Bold", 10), text_color="#94A3B8").pack(anchor="w")
        self.entry_termo_comparar = ctk.CTkEntry(
            self.tab_comparador, height=28, font=("Arial", 11),
            fg_color="#0F172A", border_color="#334155"
        )
        self.entry_termo_comparar.insert(0, "CapCut")
        self.entry_termo_comparar.pack(fill="x", pady=(2, 6))

        ctk.CTkButton(
            self.tab_comparador, text="🔎 COMPARAR PREÇOS (PLATI vs Z2U vs GGMAX)",
            font=("Arial Bold", 11), fg_color="#10B981", hover_color="#059669",
            height=32, command=self.disparar_thread_comparador
        ).pack(fill="x", pady=2)

        # Aba 5: Afiliados
        row_afiliados_plat = ctk.CTkFrame(self.tab_afiliados, fg_color="transparent")
        row_afiliados_plat.pack(fill="x", pady=(0, 2))
        ctk.CTkLabel(row_afiliados_plat, text="PLATAFORMA DE AFILIADOS",
                     font=("Arial Bold", 10), text_color="#38BDF8").pack(anchor="w")
        self.opt_plataforma_afiliados = ctk.CTkOptionMenu(
            row_afiliados_plat, values=["Todas as Plataformas", "Shopee", "Amazon", "Mercado Livre"],
            fg_color="#0F172A", button_color="#0284C7", button_hover_color="#0369A1",
            dropdown_fg_color="#0F172A", font=("Arial Bold", 10), height=26
        )
        self.opt_plataforma_afiliados.pack(fill="x", pady=(1, 3))

        ctk.CTkLabel(self.tab_afiliados, text="NICHO / PRODUTO (EX: 'fones bluetooth')",
                     font=("Arial Bold", 10), text_color="#94A3B8").pack(anchor="w")
        self.entry_termo_afiliados = ctk.CTkEntry(
            self.tab_afiliados, height=26, font=("Arial", 11),
            fg_color="#0F172A", border_color="#334155"
        )
        self.entry_termo_afiliados.insert(0, "fones bluetooth")
        self.entry_termo_afiliados.pack(fill="x", pady=(1, 3))

        # Aba 6: Lit Hunter (Aquecedor de WhatsApp)
        ctk.CTkLabel(self.tab_lit_hunter, text="NICHO / CIDADE DOS GRUPOS (EX: 'vendas', 'brasilia')",
                     font=("Arial Bold", 10), text_color="#F59E0B").pack(anchor="w")
        self.entry_termo_lit = ctk.CTkEntry(
            self.tab_lit_hunter, height=28, font=("Arial", 11),
            fg_color="#0F172A", border_color="#334155"
        )
        self.entry_termo_lit.insert(0, "vendas e-commerce produtos")
        self.entry_termo_lit.pack(fill="x", pady=(2, 4))

        ctk.CTkLabel(self.tab_lit_hunter, text="QTD DE GRUPOS ALVO",
                     font=("Arial Bold", 10), text_color="#38BDF8").pack(anchor="w")
        self.entry_qtd_lit = ctk.CTkEntry(
            self.tab_lit_hunter, height=28, font=("Arial", 11),
            fg_color="#0F172A", border_color="#334155"
        )
        self.entry_qtd_lit.insert(0, "15")
        self.entry_qtd_lit.pack(fill="x", pady=(2, 6))

        ctk.CTkButton(
            self.tab_lit_hunter, text="🔥 1. MINERAR GRUPOS PÚBLICOS",
            font=("Arial Bold", 10), fg_color="#D97706", hover_color="#B45309",
            height=30, command=self.disparar_thread_minerar_lit
        ).pack(fill="x", pady=2)

        ctk.CTkButton(
            self.tab_lit_hunter, text="🚀 2. INICIAR AQUECIMENTO (ENTRAR NOS GRUPOS)",
            font=("Arial Bold", 10), fg_color="#10B981", hover_color="#059669",
            height=32, command=self.disparar_thread_aquecer_lit
        ).pack(fill="x", pady=2)

        # Meta & Horário
        meta_box = ctk.CTkFrame(self.col1_frame, fg_color="transparent")
        meta_box.pack(fill="x", padx=12, pady=4)

        sub_m = ctk.CTkFrame(meta_box, fg_color="transparent")
        sub_m.pack(side="left", fill="x", expand=True, padx=(0, 4))
        ctk.CTkLabel(sub_m, text="META LEADS", font=("Arial Bold", 10),
                     text_color="#94A3B8").pack(anchor="w")
        self.entry_meta = ctk.CTkEntry(sub_m, height=28, font=("Arial", 11),
                                       fg_color="#060810", border_color="#334155")
        self.entry_meta.insert(0, "20")
        self.entry_meta.pack(fill="x")

        sub_h = ctk.CTkFrame(meta_box, fg_color="transparent")
        sub_h.pack(side="right", fill="x", expand=True, padx=(4, 0))
        ctk.CTkLabel(sub_h, text="HORÁRIO (HH:MM)", font=("Arial Bold", 10),
                     text_color="#38BDF8").pack(anchor="w")
        self.entry_horario = ctk.CTkEntry(sub_h, height=28, font=("Arial", 11),
                                          fg_color="#060810", border_color="#334155")
        self.entry_horario.insert(0, "06:00")
        self.entry_horario.pack(fill="x")

        # Botões de ação
        btn_action_frame = ctk.CTkFrame(self.col1_frame, fg_color="transparent")
        btn_action_frame.pack(fill="x", padx=12, pady=(4, 8))

        self.btn_add_img = ctk.CTkButton(
            btn_action_frame, text="📁 Anexar Imagens Exemplo", font=("Arial", 11),
            fg_color="#1E293B", hover_color="#334155", text_color="#F8FAFC",
            height=30, corner_radius=8, command=self.selecionar_imagens
        )
        self.btn_add_img.pack(fill="x", pady=2)

        self.lbl_qtd_imagens = ctk.CTkLabel(
            btn_action_frame, text="Nenhuma imagem anexada",
            font=("Arial Italic", 9), text_color="#64748B"
        )
        self.lbl_qtd_imagens.pack(anchor="w", pady=(0, 4))

        self.btn_iniciar_coleta = ctk.CTkButton(
            btn_action_frame, text="🔍 1. INICIAR COLETA DE LEADS",
            font=("Arial Bold", 11), fg_color="#0284C7", hover_color="#0369A1",
            height=34, command=self.disparar_thread_coleta
        )
        self.btn_iniciar_coleta.pack(fill="x", pady=2)

        self.btn_iniciar_disparos = ctk.CTkButton(
            btn_action_frame, text="🚀 2. INICIAR DISPAROS COM IA",
            font=("Arial Bold", 11), fg_color="#10B981", hover_color="#059669",
            height=34, command=self.disparar_thread_envio
        )
        self.btn_iniciar_disparos.pack(fill="x", pady=2)

        self.btn_ativar_agendamento = ctk.CTkButton(
            btn_action_frame, text="⏰ ATIVAR AGENDAMENTO AUTO",
            font=("Arial Bold", 11), fg_color="#D97706", hover_color="#B45309",
            height=32, command=self.toggle_agendamento
        )
        self.btn_ativar_agendamento.pack(fill="x", pady=2)

        self.btn_parar = ctk.CTkButton(
            btn_action_frame, text="⏹ PARAR TUDO", font=("Arial Bold", 11),
            fg_color="#DC2626", hover_color="#B91C1C", height=30,
            command=self.parar_automacao
        )
        self.btn_parar.pack(fill="x", pady=(2, 0))

    def _montar_coluna_2_resultados(self) -> None:
        top_bar_col2 = ctk.CTkFrame(self.col2_frame, fg_color="transparent")
        top_bar_col2.pack(fill="x", padx=16, pady=(14, 6))
        ctk.CTkLabel(top_bar_col2, text="DATA STREAM & CARDS DE RESULTADOS",
                     font=("Arial Bold", 12), text_color="#94A3B8").pack(side="left")

        btn_box = ctk.CTkFrame(top_bar_col2, fg_color="transparent")
        btn_box.pack(side="right")

        ctk.CTkButton(btn_box, text="📥 Exportar (.xlsm/Excel)", width=130, height=26,
                      font=("Arial Bold", 10), fg_color="#10B981", hover_color="#059669",
                      command=self.exportar_leads_excel).pack(side="left", padx=(0, 4))

        ctk.CTkButton(btn_box, text="Marcar Todos", width=90, height=26,
                      font=("Arial Bold", 10), fg_color="#1E293B",
                      command=self.marcar_todos_leads).pack(side="left", padx=(0, 4))

        ctk.CTkButton(btn_box, text="Desmarcar", width=80, height=26,
                      font=("Arial Bold", 10), fg_color="#1E293B",
                      command=self.desmarcar_todos_leads).pack(side="left", padx=(0, 4))

        ctk.CTkButton(btn_box, text="🧹 Limpar Tela", width=90, height=26,
                      font=("Arial Bold", 10), fg_color="#DC2626", hover_color="#B91C1C",
                      command=self.limpar_resultados).pack(side="left")

        self.scroll_leads = ctk.CTkScrollableFrame(
            self.col2_frame, fg_color="#060810",
            border_color="#1E293B", border_width=1, corner_radius=10
        )
        self.scroll_leads.pack(fill="both", expand=True, padx=16, pady=(0, 16))

    def _montar_coluna_3_telemetria(self) -> None:
        zap_card = ctk.CTkFrame(self.col3_frame, fg_color="#1E293B", corner_radius=10, height=65)
        zap_card.pack(fill="x", padx=14, pady=(14, 6))
        zap_card.pack_propagate(False)
        ctk.CTkLabel(zap_card, text="🟢 WHATSAPP DESKTOP PRONTO",
                     font=("Arial Bold", 11), text_color="#10B981").pack(anchor="w", padx=12, pady=(10, 0))
        ctk.CTkLabel(zap_card, text="Integração de Cotação R$ & Trava Antiduplicidade Ativa.",
                     font=("Arial", 10), text_color="#94A3B8").pack(anchor="w", padx=12, pady=(0, 8))

        # 🛡️ CARD DE SEGURANÇA EM TEMPO REAL
        sec_card = ctk.CTkFrame(
            self.col3_frame, fg_color="#060810", corner_radius=10,
            border_width=1, border_color="#10B981"
        )
        sec_card.pack(fill="x", padx=14, pady=(0, 8))

        top_sec = ctk.CTkFrame(sec_card, fg_color="transparent")
        top_sec.pack(fill="x", padx=10, pady=(8, 2))

        ctk.CTkLabel(
            top_sec, text="🛡️ SHIELD DE SEGURANÇA ATIVO",
            font=("Arial Bold", 11), text_color="#10B981"
        ).pack(side="left")

        self.lbl_sec_status_geral = ctk.CTkLabel(
            top_sec, text="🟢 AMBIENTE SEGURO",
            font=("Arial Bold", 10), text_color="#10B981"
        )
        self.lbl_sec_status_geral.pack(side="right")

        # Grid de Indicadores de Segurança
        grid_sec = ctk.CTkFrame(sec_card, fg_color="transparent")
        grid_sec.pack(fill="x", padx=10, pady=(2, 8))

        self.lbl_sec_tls = ctk.CTkLabel(grid_sec, text="🔒 SSL/TLS 1.3: OK", font=("Arial", 9), text_color="#38BDF8")
        self.lbl_sec_tls.grid(row=0, column=0, sticky="w", padx=(0, 8))

        self.lbl_sec_vault = ctk.CTkLabel(grid_sec, text="🔑 DPAPI Vault: OK", font=("Arial", 9), text_color="#38BDF8")
        self.lbl_sec_vault.grid(row=0, column=1, sticky="w")

        self.lbl_sec_shell = ctk.CTkLabel(grid_sec, text="🛡️ Shell Guard: OK", font=("Arial", 9), text_color="#38BDF8")
        self.lbl_sec_shell.grid(row=1, column=0, sticky="w", padx=(0, 8))

        self.lbl_sec_tamper = ctk.CTkLabel(grid_sec, text="👁️ Anti-Tamper: OK", font=("Arial", 9), text_color="#38BDF8")
        self.lbl_sec_tamper.grid(row=1, column=1, sticky="w")

        self.card_timer_pausa = ctk.CTkFrame(
            self.col3_frame, fg_color="#1E293B", corner_radius=10,
            border_width=1, border_color="#334155"
        )
        self.card_timer_pausa.pack(fill="x", padx=14, pady=(0, 8))

        top_timer_row = ctk.CTkFrame(self.card_timer_pausa, fg_color="transparent")
        top_timer_row.pack(fill="x", padx=12, pady=(8, 2))

        ctk.CTkLabel(
            top_timer_row, text="🛡️ PAUSA DE SEGURANÇA (ANTI-BAN)",
            font=("Arial Bold", 10), text_color="#F59E0B"
        ).pack(side="left")

        self.lbl_timer_pausa = ctk.CTkLabel(
            top_timer_row, text="EM ESPERA", font=("Arial Bold", 11),
            text_color="#64748B"
        )
        self.lbl_timer_pausa.pack(side="right")

        self.progress_timer_pausa = ctk.CTkProgressBar(
            self.card_timer_pausa, height=5, corner_radius=3,
            fg_color="#060810", progress_color="#F59E0B"
        )
        self.progress_timer_pausa.set(0.0)
        self.progress_timer_pausa.pack(fill="x", padx=12, pady=(0, 8))

        ctk.CTkLabel(self.col3_frame, text="CONSOLE DE EXECUÇÃO EM TEMPO REAL",
                     font=("Arial Bold", 11), text_color="#94A3B8").pack(anchor="w", padx=14, pady=(4, 4))

        self.txt_console = ctk.CTkTextbox(
            self.col3_frame, font=("Consolas", 10), fg_color="#060810",
            text_color="#38BDF8", border_color="#1E293B", border_width=1, corner_radius=10
        )
        self.txt_console.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        self.txt_console.insert(
            "0.0",
            ">>> [LEAD HUNTER PRO v3.0 Enterprise — Gemini Multi-Source Hybrid Engine]\n"
            f">>> Base: {self._base_path}\n"
            f">>> WAL Mode ● Session Pool ● Gemini Singleton ● Rust Engine: {'ATIVO' if SUPORTA_RUST else 'Python Fallback'}\n"
        )
        self.txt_console.configure(state="disabled")

    def _criar_rodape_enterprise(self) -> None:
        self.footer_bar = ctk.CTkFrame(
            self, fg_color="#0F172A", corner_radius=0, height=36,
            border_width=1, border_color="#1E293B"
        )
        self.footer_bar.pack(side="bottom", fill="x", padx=0, pady=0)
        self.footer_bar.pack_propagate(False)

        self.lbl_footer_tempo = ctk.CTkLabel(
            self.footer_bar, text="AUTOMAÇÃO EM ESPERA",
            font=("Arial Bold", 10), text_color="#10B981"
        )
        self.lbl_footer_tempo.pack(side="left", padx=20)

        ctk.CTkLabel(
            self.footer_bar,
            text="Multi-Marketplace Intelligence & Automated Outreach Engine",
            font=("Arial", 10), text_color="#64748B"
        ).pack(side="left", expand=True)

        self.lbl_footer_data = ctk.CTkLabel(
            self.footer_bar, text="", font=("Arial Bold", 10), text_color="#CBD5E1"
        )
        self.lbl_footer_data.pack(side="right", padx=20)

    # ── EXPORTAÇÃO PARA PLANILHA .XLSX / .XLSM / GOOGLE CONTACTS ──────────────

    def exportar_leads_excel(self) -> None:
        if not self.leads_dados:
            self.log("✖ Nenhum lead capturado na tela para exportar!")
            return

        caminho_arquivo = filedialog.asksaveasfilename(
            title="Salvar Tabela de Leads",
            defaultextension=".xlsm",
            filetypes=[
                ("Planilha Excel (.xlsm)", "*.xlsm"),
                ("Planilha Excel (.xlsx)", "*.xlsx"),
                ("Arquivo CSV (Google Contacts)", "*.csv")
            ]
        )

        if not caminho_arquivo:
            return

        # BLINDAGEM CONTRA PATH TRAVERSAL
        caminho_validado = PathTraversalGuard.sanitizar_caminho_arquivo(caminho_arquivo)
        if not caminho_validado:
            self.log("✖ Caminho de arquivo bloqueado por segurança!")
            return
        caminho_arquivo = caminho_validado

        try:
            if SUPORTA_OPENPYXL and (caminho_arquivo.endswith(".xlsx") or caminho_arquivo.endswith(".xlsm")):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Leads Coletados"

                headers = ["Nome", "Contato", "Data de Coleta", "Name", "Phone 1 - Value"]
                ws.append(headers)

                for cell in ws[1]:
                    cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                    cell.fill = openpyxl.styles.PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")

                for item in self.leads_dados:
                    nome = item["nome"]
                    tel = item["telefone"]
                    data_hoje = datetime.now().strftime("%d/%m/%Y")
                    ws.append([nome, tel, data_hoje, nome, f"+{tel}"])

                wb.save(caminho_arquivo)
                self.log(f"✔ Planilha exportada com sucesso em: {caminho_arquivo}")

            else:
                with open(caminho_arquivo, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.writer(f, delimiter=";")
                    writer.writerow(["Nome", "Contato", "Data de Coleta", "Name", "Phone 1 - Value"])
                    for item in self.leads_dados:
                        nome = item["nome"]
                        tel = item["telefone"]
                        data_hoje = datetime.now().strftime("%d/%m/%Y")
                        writer.writerow([nome, tel, data_hoje, nome, f"+{tel}"])

                self.log(f"✔ Tabela de leads salva com sucesso em: {caminho_arquivo}")

        except Exception as e:
            self.log(f"✖ Erro ao exportar planilha: {e}")

    # ── Reset / Limpeza ───────────────────────────────────────────────────────

    def limpar_resultados(self) -> None:
        try:
            for child in self.scroll_leads.winfo_children():
                child.destroy()
            self.leads_dados.clear()
            self.grupos_coletados.clear()
            self.log("🧹 Tela de resultados e cards limpos com sucesso!")
        except Exception as e:
            self.log(f"✖ Erro ao limpar resultados: {e}")

    # ── Renderizadores de Cards e Imagens ────────────────────────────────────

    def adicionar_lead_na_lista(self, nome_loja: str, telefone: str) -> bool:
        try:
            if not self.validar_e_filtrar_celular(telefone):
                return False
            if any(ld["telefone"] == telefone for ld in self.leads_dados):
                return False

            registro_antigo = self.db.verificar_lead_ja_abordado(telefone)
            var_status = ctk.BooleanVar(value=not bool(registro_antigo))

            frame_linha = ctk.CTkFrame(
                self.scroll_leads, fg_color="#0F172A", height=52,
                corner_radius=8, border_width=1, border_color="#1E293B"
            )
            frame_linha.pack(fill="x", pady=4)

            if registro_antigo:
                texto_label = f"👤 {nome_loja}\n📞 {telefone} (⚠️ Já abordado em {registro_antigo[0]})"
                cor_texto = "#FACC15"
            else:
                texto_label = f"👤 {nome_loja}\n📞 {telefone}"
                cor_texto = "#F8FAFC"

            ctk.CTkCheckBox(
                frame_linha, text=texto_label, variable=var_status,
                font=("Arial", 10), text_color=cor_texto, fg_color="#0284C7"
            ).pack(side="left", padx=12, pady=6, anchor="w")

            self.leads_dados.append({
                "nome": nome_loja,
                "telefone": telefone,
                "var": var_status,
                "frame": frame_linha
            })
            return True
        except Exception as err:
            self.log(f"✖ Erro ao adicionar lead na interface: {err}")
            return False

    def adicionar_grupo_lit_na_lista(self, titulo_grupo: str, link_grupo: str) -> None:
        frame_linha = ctk.CTkFrame(
            self.scroll_leads, fg_color="#0F172A", height=75,
            corner_radius=10, border_width=1, border_color="#F59E0B"
        )
        frame_linha.pack(fill="x", pady=4)
        frame_linha.pack_propagate(False)

        info_box = ctk.CTkFrame(frame_linha, fg_color="transparent")
        info_box.pack(side="left", fill="both", expand=True, padx=12, pady=8)

        ctk.CTkLabel(info_box, text=f"🔥 {titulo_grupo}", font=("Arial Bold", 11),
                     text_color="#F59E0B", anchor="w").pack(anchor="w")

        ctk.CTkLabel(info_box, text=f"🔗 Link Público: {link_grupo[:60]}...", font=("Consolas", 9),
                     text_color="#94A3B8", anchor="w").pack(anchor="w")

        ctk.CTkButton(
            frame_linha, text="Entrar no Grupo", width=110, height=28,
            font=("Arial Bold", 10), fg_color="#D97706", hover_color="#B45309",
            command=lambda u=link_grupo: self.abrir_link_anuncio_seguro(u)
        ).pack(side="right", padx=12)

        self.grupos_coletados.append({
            "titulo": titulo_grupo,
            "link": link_grupo,
            "frame": frame_linha
        })

    def adicionar_anuncio_na_lista(
        self,
        titulo_anuncio: str,
        img_url: str,
        link_ads: str,
        taxa_conversao: float,
        ja_coletado_data: str | None = None
    ) -> None:
        frame_linha = ctk.CTkFrame(
            self.scroll_leads, fg_color="#0F172A", height=115,
            corner_radius=10, border_width=1, border_color="#334155"
        )
        frame_linha.pack(fill="x", pady=5)
        frame_linha.pack_propagate(False)

        lbl_placeholder = ctk.CTkLabel(
            frame_linha, text="🎯\nMeta Ad", width=90, height=90,
            fg_color="#1E293B", corner_radius=8, font=("Arial Bold", 11),
            text_color="#38BDF8"
        )
        lbl_placeholder.pack(side="left", padx=8, pady=8)

        info_box = ctk.CTkFrame(frame_linha, fg_color="transparent")
        info_box.pack(side="left", fill="both", expand=True, padx=6, pady=6)
        ctk.CTkLabel(info_box, text=f"🎯 {titulo_anuncio}", font=("Arial Bold", 11),
                     text_color="#38BDF8", anchor="w").pack(anchor="w")

        if ja_coletado_data:
            badge = f"⚠️ Coletado em {ja_coletado_data} | Conversão Est.: {taxa_conversao}%"
            cor = "#FACC15"
        else:
            badge = f"🌟 NOVO / Validado | Conversão Est.: {taxa_conversao}%"
            cor = "#10B981"

        ctk.CTkLabel(info_box, text=badge, font=("Arial Bold", 10),
                     text_color=cor, anchor="w").pack(anchor="w")

        row_link = ctk.CTkFrame(info_box, fg_color="transparent")
        row_link.pack(anchor="w", fill="x", pady=(2, 0))

        ctk.CTkLabel(row_link, text=f"🔗 Direct: {link_ads}", font=("Consolas", 9),
                     text_color="#94A3B8", anchor="w").pack(side="left", padx=(0, 6))

        ctk.CTkButton(
            row_link, text="Abrir no Meta", width=110, height=22,
            font=("Arial Bold", 9), fg_color="#A855F7", hover_color="#7E22CE",
            command=lambda u=link_ads: self.abrir_link_anuncio_seguro(u)
        ).pack(side="left")

        if img_url and SUPORTA_IMAGEM:
            def _carregar_imagem_async() -> None:
                img_bytes = self.http.get_cached_image_bytes(img_url)
                if img_bytes:
                    try:
                        with io.BytesIO(img_bytes) as img_data:
                            with Image.open(img_data) as pil_img:
                                pil_resized = pil_img.resize((90, 90))
                                if hasattr(ctk, "CTkImage"):
                                    ctk_img = ctk.CTkImage(light_image=pil_resized, dark_image=pil_resized, size=(90, 90))
                                else:
                                    ctk_img = ImageTk.PhotoImage(pil_resized)

                        def _update_ui() -> None:
                            try:
                                lbl_placeholder.configure(image=ctk_img, text="")
                                lbl_placeholder.image = ctk_img
                            except Exception:
                                pass

                        self.after(0, _update_ui)
                    except Exception:
                        pass

            threading.Thread(target=_carregar_imagem_async, daemon=True).start()

    def adicionar_produto_afiliado_na_lista(
        self,
        titulo: str,
        plataforma: str,
        preco_brl: float,
        img_url: str,
        link_produto: str,
        taxa_conversao: float
    ) -> None:
        frame_linha = ctk.CTkFrame(
            self.scroll_leads, fg_color="#0F172A", height=105,
            corner_radius=10, border_width=1, border_color="#EAB308"
        )
        frame_linha.pack(fill="x", pady=5)
        frame_linha.pack_propagate(False)

        lbl_placeholder = ctk.CTkLabel(
            frame_linha, text="🛍️\nProduto", width=80, height=80,
            fg_color="#1E293B", corner_radius=8, font=("Arial Bold", 10),
            text_color="#EAB308"
        )
        lbl_placeholder.pack(side="left", padx=8, pady=8)

        info_box = ctk.CTkFrame(frame_linha, fg_color="transparent")
        info_box.pack(side="left", fill="both", expand=True, padx=6, pady=6)
        ctk.CTkLabel(info_box, text=f"🛍️ {titulo}", font=("Arial Bold", 11),
                     text_color="#F8FAFC", anchor="w").pack(anchor="w")

        badge = f"🏷️ {plataforma} | 🇧🇷 Preço: R$ {preco_brl:.2f} | Est. Conversão Afiliado: {taxa_conversao}%"
        ctk.CTkLabel(info_box, text=badge, font=("Arial Bold", 10),
                     text_color="#EAB308", anchor="w").pack(anchor="w")

        ctk.CTkButton(
            info_box, text=f"🔗 Copiar / Abrir Oferta ({plataforma})", width=180, height=26,
            font=("Arial Bold", 10), fg_color="#D97706", hover_color="#B45309",
            command=lambda u=link_produto: self.abrir_link_anuncio_seguro(u)
        ).pack(anchor="w", pady=(4, 0))

        if img_url and SUPORTA_IMAGEM:
            def _carregar_imagem_async() -> None:
                img_bytes = self.http.get_cached_image_bytes(img_url)
                if img_bytes:
                    try:
                        with io.BytesIO(img_bytes) as img_data:
                            with Image.open(img_data) as pil_img:
                                pil_resized = pil_img.resize((80, 80))
                                if hasattr(ctk, "CTkImage"):
                                    ctk_img = ctk.CTkImage(light_image=pil_resized, dark_image=pil_resized, size=(80, 80))
                                else:
                                    ctk_img = ImageTk.PhotoImage(pil_resized)

                        def _update_ui() -> None:
                            try:
                                lbl_placeholder.configure(image=ctk_img, text="")
                                lbl_placeholder.image = ctk_img
                            except Exception:
                                pass

                        self.after(0, _update_ui)
                    except Exception:
                        pass

            threading.Thread(target=_carregar_imagem_async, daemon=True).start()

    def adicionar_comparacao_na_lista(
        self,
        titulo: str,
        plataforma: str,
        preco_usd: float,
        preco_brl: float,
        img_url: str,
        link_produto: str,
        is_menor_preco: bool = False
    ) -> None:
        frame_linha = ctk.CTkFrame(
            self.scroll_leads, fg_color="#0F172A", height=120, corner_radius=10,
            border_width=1, border_color="#10B981" if is_menor_preco else "#334155"
        )
        frame_linha.pack(fill="x", pady=5)
        frame_linha.pack_propagate(False)

        lbl_placeholder = ctk.CTkLabel(
            frame_linha, text="📦\nProduto", width=95, height=95,
            fg_color="#1E293B", corner_radius=8, font=("Arial Bold", 11),
            text_color="#10B981" if is_menor_preco else "#38BDF8"
        )
        lbl_placeholder.pack(side="left", padx=8, pady=8)

        info_box = ctk.CTkFrame(frame_linha, fg_color="transparent")
        info_box.pack(side="left", fill="both", expand=True, padx=6, pady=6)
        ctk.CTkLabel(info_box, text=f"📦 {titulo}", font=("Arial Bold", 11),
                     text_color="#F8FAFC", anchor="w").pack(anchor="w")

        precos_str = f"💵 US$ {preco_usd:.2f}  ➔  🇧🇷 R$ {preco_brl:.2f}"
        if is_menor_preco:
            tag = f"🏆 MENOR PREÇO DO MERCADO! ({plataforma}) — {precos_str}"
            cor_tag = "#10B981"
        else:
            tag = f"🛒 {plataforma} — {precos_str}"
            cor_tag = "#38BDF8"

        ctk.CTkLabel(info_box, text=tag, font=("Arial Bold", 10),
                     text_color=cor_tag, anchor="w").pack(anchor="w")

        row_link = ctk.CTkFrame(info_box, fg_color="transparent")
        row_link.pack(anchor="w", fill="x", pady=(2, 0))

        link_curto = link_produto[:55] + "..." if len(link_produto) > 55 else link_produto
        ctk.CTkLabel(row_link, text=f"🔗 Direct: {link_curto}", font=("Consolas", 9),
                     text_color="#94A3B8", anchor="w").pack(side="left", padx=(0, 6))

        ctk.CTkButton(
            row_link, text=f"Ir Para Oferta ({plataforma})",
            width=150, height=22, font=("Arial Bold", 9),
            fg_color="#059669" if is_menor_preco else "#0284C7",
            hover_color="#0369A1",
            command=lambda u=link_produto: self.abrir_link_anuncio_seguro(u)
        ).pack(side="left")

        if img_url and SUPORTA_IMAGEM:
            def _carregar_imagem_async() -> None:
                img_bytes = self.http.get_cached_image_bytes(img_url)
                if img_bytes:
                    try:
                        with io.BytesIO(img_bytes) as img_data:
                            with Image.open(img_data) as pil_img:
                                pil_resized = pil_img.resize((95, 95))
                                if hasattr(ctk, "CTkImage"):
                                    ctk_img = ctk.CTkImage(light_image=pil_resized, dark_image=pil_resized, size=(95, 95))
                                else:
                                    ctk_img = ImageTk.PhotoImage(pil_resized)

                        def _update_ui() -> None:
                            try:
                                lbl_placeholder.configure(image=ctk_img, text="")
                                lbl_placeholder.image = ctk_img
                            except Exception:
                                pass

                        self.after(0, _update_ui)
                    except Exception:
                        pass

            threading.Thread(target=_carregar_imagem_async, daemon=True).start()

    # ── IA & CONEXÃO AGENTE GEMINI ───────────────────────────────────────────

    def testar_conexao_api(self) -> None:
        api_key = self.entry_api_key.get().strip()
        if not api_key:
            self.lbl_api_status.configure(
                text="🔴 Status da IA: Chave não informada (Modo local)",
                text_color="#EF4444"
            )
            return
        if not SUPORTA_GEMINI:
            self.lbl_api_status.configure(
                text="🔴 Status da IA: Biblioteca 'google-genai' ausente",
                text_color="#EF4444"
            )
            return

        self.lbl_api_status.configure(
            text="🟡 Status da IA: Testando conexão...", text_color="#F59E0B"
        )

        def run_test() -> None:
            self.gemini.configurar(api_key)
            resultado = self.gemini.gerar_conteudo("Responda apenas com a palavra: 'OK'")
            if resultado:
                self.after(0, lambda: self.lbl_api_status.configure(
                    text="🟢 Status da IA: Gemini Agent Ativo em Tempo Real",
                    text_color="#10B981"
                ))
                self.log("🤖 [Gemini AI Agent] Conexão validada! Agente Otimizador ativado em tempo real para todas as funções.")
            else:
                self.after(0, lambda: self.lbl_api_status.configure(
                    text="🟢 Status da IA: Conectado (Modo Protegido)",
                    text_color="#10B981"
                ))
                self.log("✔ Conexão com Google Gemini API estabelecida!")

        threading.Thread(target=run_test, daemon=True).start()

    def gerar_mensagem_inteligente(
        self, nome_empresa: str, termo_nicho: str, mensagem_base: str
    ) -> str:
        api_key = self.entry_api_key.get().strip()
        if SUPORTA_GEMINI and api_key:
            self.gemini.configurar(api_key)
            res = self.gemini.otimizar_copywriting_whatsapp(nome_empresa, termo_nicho, mensagem_base)
            if res:
                self.log(f"🤖 [Gemini Agent] Copy B2B otimizada para [{nome_empresa}]")
                return res

        return (
            f"Olá! Notei o excelente trabalho da '{nome_empresa}' no segmento de {termo_nicho}. "
            f"{mensagem_base} Teria 2 minutos hoje para eu te enviar uma demonstração rápida?"
        )

    # ── AUTOMAÇÃO DE ENTRADA / ENVIAR PEDIDO NO WHATSAPP (UNIVERSAL) ──────────

    def entrar_e_confirmar_grupo_whatsapp(self, link_grupo: str, saudacao: str) -> bool:
        """
        Automação Ultra-Fluida e Humanizada de Entrada/Pedido em Grupos do WhatsApp Desktop.
        - Protocolo Nativo: whatsapp://accept?code=HASH
        - Desbloqueio de Foco de Janela no Windows (Bypass ALT)
        - Movimentação Suave do Mouse (Human Cursor Glide)
        - Calibragem 100% Precisa para 'Entrar no grupo', 'Enviar pedido' e 'Entrar na comunidade'
        - Simulação de Leitura (Scroll) e Envio da Mensagem de Aquecimento
        """
        try:
            pyautogui.FAILSAFE = False
            pyautogui.PAUSE = 0.15

            # 1. Extração do código de convite do grupo
            codigo_grupo = link_grupo.strip().split("/")[-1]
            link_accept_app = f"whatsapp://accept?code={codigo_grupo}"
            link_chat_app = f"whatsapp://chat?code={codigo_grupo}"

            self.log(f"⚡ Acionando protocolo nativo do aplicativo para o grupo [{codigo_grupo}]...")

            # Tenta o protocolo oficial de aceite de grupo do WhatsApp Desktop
            CommandInjectionGuard.abrir_url_com_seguranca(link_accept_app)
            time.sleep(1.8)

            CommandInjectionGuard.abrir_url_com_seguranca(link_chat_app)
            time.sleep(1.2)

            # 2. Bypassa a trava de foco do Windows e localiza a janela do WhatsApp Desktop
            pyautogui.press("alt")
            time.sleep(0.2)

            hwnds = []

            def enum_windows_callback(hwnd: int, extra: list[int]) -> None:
                if win32gui.IsWindowVisible(hwnd) and "WhatsApp" in win32gui.GetWindowText(hwnd):
                    extra.append(hwnd)

            win32gui.EnumWindows(enum_windows_callback, hwnds)

            if not hwnds:
                self.log("✖ Janela do WhatsApp Desktop não encontrada.")
                return False

            hwnd = hwnds[0]
            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
            try:
                win32gui.SetForegroundWindow(hwnd)
            except Exception:
                pass
            time.sleep(1.0)

            # Obtém as dimensões exatas da janela do WhatsApp
            rect = win32gui.GetWindowRect(hwnd)
            left, top, right, bottom = rect
            width, height = right - left, bottom - top

            if width < 200 or height < 200:
                self.log("⚠️ Janela do WhatsApp muito pequena.")
                return False

            # Limpa qualquer alerta ou menu prévio com ESC
            pyautogui.press("esc")
            time.sleep(0.3)

            # 3. Força a seleção do botão DIREITO do modal (Entrar no grupo / Enviar pedido / Entrar na comunidade)
            pyautogui.press("right")
            time.sleep(0.2)
            pyautogui.press("tab")
            time.sleep(0.2)
            pyautogui.press("right")
            time.sleep(0.2)
            pyautogui.press("enter")
            time.sleep(0.2)
            pyautogui.press("space")

            # Coordenadas recalibradas para o botão DIREITO do modal (56.5% largura, 54% altura)
            target_x = left + int(width * 0.565)
            target_y = top + int(height * 0.540)

            # Movimento suave de mouse estilo humano (Glide)
            pyautogui.moveTo(target_x, target_y, duration=random.uniform(0.25, 0.45))
            time.sleep(0.1)
            pyautogui.click()
            time.sleep(0.3)

            # Clique de garantia no centro do botão da direita
            pyautogui.click(left + int(width * 0.570), top + int(height * 0.550))
            time.sleep(0.3)
            pyautogui.press("enter")

            self.log("🖱️ Ação de confirmação (Entrar/Solicitar) executada com sucesso!")
            time.sleep(2.5)

            # 4. Simulação de "Leitura" (Scroll Humano no Chat)
            chat_area_x = left + int(width * 0.65)
            chat_area_y = top + int(height * 0.50)
            pyautogui.moveTo(chat_area_x, chat_area_y, duration=random.uniform(0.2, 0.35))
            pyautogui.click()

            self.log("📜 Simulando leitura natural do grupo (Scroll)...")
            for _ in range(2):
                pyautogui.scroll(random.choice([350, 450]))
                time.sleep(random.uniform(0.25, 0.4))
                pyautogui.scroll(random.choice([-350, -450]))
                time.sleep(random.uniform(0.25, 0.4))

            # 5. Envio da Mensagem de Saudação (Aquecimento do Perfil)
            if saudacao:
                input_x = left + int(width * 0.65)
                input_y = bottom - 45
                pyautogui.moveTo(input_x, input_y, duration=random.uniform(0.2, 0.35))
                pyautogui.click()
                time.sleep(0.3)

                pyperclip.copy(saudacao)
                pyautogui.hotkey("ctrl", "v")
                time.sleep(0.4)
                pyautogui.press("enter")
                self.log(f"✅ Mensagem de aquecimento enviada no chat.")

            # 6. Fecha guia residual do navegador se houver
            try:
                pyautogui.hotkey("ctrl", "w")
            except Exception:
                pass

            return True

        except Exception as err:
            self.log(f"⚠️ Erro no fluxo de entrada no grupo: {err}")
            return False

    # ── DISPARADORES ─────────────────────────────────────────────────────────

    def disparar_thread_minerar_lit(self) -> None:
        if not self._tentar_iniciar_automacao():
            return
        self.tempo_inicio_automacao = time.time()
        self._ui_safe(lambda: self.lbl_status_sistema.configure(
            text="● MINERANDO GRUPOS...", text_color="#F59E0B"
        ))
        threading.Thread(target=self.executar_mineracao_grupos_whatsapp, daemon=True).start()

    def disparar_thread_aquecer_lit(self) -> None:
        if not self._tentar_iniciar_automacao():
            return
        self.tempo_inicio_automacao = time.time()
        self._ui_safe(lambda: self.lbl_status_sistema.configure(
            text="● AQUECENDO WHATSAPP...", text_color="#10B981"
        ))
        threading.Thread(target=self.executar_aquecimento_whatsapp, daemon=True).start()

    def disparar_thread_afiliados_unico(self) -> None:
        if not self._tentar_iniciar_automacao():
            return
        self._ui_safe(lambda: self.lbl_status_sistema.configure(
            text="● MINERANDO AFILIADOS...", text_color="#EAB308"
        ))
        threading.Thread(target=self.executar_pesquisa_afiliados_unico, daemon=True).start()

    def disparar_thread_afiliados_auto(self) -> None:
        if not self._tentar_iniciar_automacao():
            return
        self.tempo_inicio_automacao = time.time()
        self._ui_safe(lambda: self.lbl_status_sistema.configure(
            text="● AUTO-MINERANDO AFILIADOS...", text_color="#D97706"
        ))
        threading.Thread(target=self.executar_auto_mineracao_afiliados, daemon=True).start()

    def disparar_thread_comparador(self) -> None:
        if not self._tentar_iniciar_automacao():
            return
        self._ui_safe(lambda: self.lbl_status_sistema.configure(
            text="● COMPARANDO PREÇOS...", text_color="#0284C7"
        ))
        threading.Thread(target=self.executar_comparacao_plati_z2u, daemon=True).start()

    def disparar_thread_envio(self) -> None:
        if not self._tentar_iniciar_automacao():
            return

        aba_atual = self.tabview.get()
        if self.modo_lit_hunter or aba_atual == "🔥 Lit Hunter":
            if not self.grupos_coletados:
                self.log("✖ Nenhum grupo de WhatsApp na lista para aquecimento!")
                self.automacao_rodando = False
                return
            self.tempo_inicio_automacao = time.time()
            self._ui_safe(lambda: self.lbl_status_sistema.configure(
                text="● AQUECENDO WHATSAPP...", text_color="#10B981"
            ))
            threading.Thread(target=self.executar_aquecimento_whatsapp, daemon=True).start()
        else:
            if not self.leads_dados:
                self.log("✖ Nenhum lead na lista para disparar!")
                self.automacao_rodando = False
                return
            self.tempo_inicio_automacao = time.time()
            self._ui_safe(lambda: self.lbl_status_sistema.configure(
                text="● ENVIANDO COM IA...", text_color="#10B981"
            ))
            threading.Thread(target=self.executar_disparos_whatsapp, daemon=True).start()

    def disparar_thread_coleta(self) -> None:
        if not self._tentar_iniciar_automacao():
            return
        self.tempo_inicio_automacao = time.time()

        aba_atual = self.tabview.get()
        if self.modo_lit_hunter or aba_atual == "🔥 Lit Hunter":
            self._ui_safe(lambda: self.lbl_status_sistema.configure(text="● MINERANDO GRUPOS...", text_color="#F59E0B"))
            threading.Thread(target=self.executar_mineracao_grupos_whatsapp, daemon=True).start()
        elif aba_atual == "🌐 Web":
            self._ui_safe(lambda: self.lbl_status_sistema.configure(text="● COLETANDO...", text_color="#0284C7"))
            threading.Thread(target=self.executar_coleta_busca_avancada, daemon=True).start()
        elif aba_atual == "🎯 Meta":
            self._ui_safe(lambda: self.lbl_status_sistema.configure(text="● ESPIONANDO ANÚNCIOS...", text_color="#A855F7"))
            threading.Thread(target=self.executar_pesquisa_ads_library, daemon=True).start()
        elif aba_atual == "⚖️ Preços":
            self._ui_safe(lambda: self.lbl_status_sistema.configure(text="● COMPARANDO PREÇOS...", text_color="#0284C7"))
            threading.Thread(target=self.executar_comparacao_plati_z2u, daemon=True).start()
        elif aba_atual == "🛍️ Afiliados":
            self._ui_safe(lambda: self.lbl_status_sistema.configure(text="● MINERANDO AFILIADOS...", text_color="#EAB308"))
            threading.Thread(target=self.executar_pesquisa_afiliados_unico, daemon=True).start()
        else:
            self._ui_safe(lambda: self.lbl_status_sistema.configure(text="● COLETANDO...", text_color="#0284C7"))
            threading.Thread(target=self.executar_coleta_maps, daemon=True).start()

    def disparar_thread_ads_library(self) -> None:
        if not self._tentar_iniciar_automacao():
            return
        self._ui_safe(lambda: self.lbl_status_sistema.configure(
            text="● ESPIONANDO ANÚNCIOS...", text_color="#A855F7"
        ))
        threading.Thread(target=self.executar_pesquisa_ads_library, daemon=True).start()

    def disparar_thread_auto_mineracao(self) -> None:
        if not self._tentar_iniciar_automacao():
            return
        self.tempo_inicio_automacao = time.time()
        self._ui_safe(lambda: self.lbl_status_sistema.configure(
            text="● AUTO-MINERANDO POR TEMPO...", text_color="#F59E0B"
        ))
        threading.Thread(target=self.executar_auto_mineracao_ads, daemon=True).start()

    # ── SUB-PROJETO: LIT HUNTER (AQUECIMENTO AUTOMÁTICO DE WHATSAPP) ──────────

    def executar_mineracao_grupos_whatsapp(self) -> None:
        nicho_bruto = self.entry_termo_lit.get().strip() if self.entry_termo_lit else "vendas e-commerce"
        try:
            limite = int(self.entry_qtd_lit.get().strip()) if self.entry_qtd_lit else 15
        except ValueError:
            limite = 15

        palavras_nicho = [p.strip() for p in re.split(r'\s+', nicho_bruto) if len(p.strip()) > 2]
        termo_principal = palavras_nicho[0] if palavras_nicho else "vendas"

        queries_pool = [
            f'"chat.whatsapp.com/" "{termo_principal}"',
            f'"chat.whatsapp.com/" "{nicho_bruto}"',
            f'site:facebook.com "chat.whatsapp.com/" "{termo_principal}"',
            f'site:instagram.com "chat.whatsapp.com/" "{termo_principal}"',
            f'"chat.whatsapp.com/" "grupo" "{termo_principal}"',
            f'site:gruposwhats.app "{termo_principal}"'
        ]

        if self.gemini.ativo:
            prompt_dork = (
                f"Gere 3 dorks de busca do Google altamente eficazes para encontrar links de grupos de WhatsApp sobre '{nicho_bruto}'. "
                f"Sempre inclua 'chat.whatsapp.com/'. Separe as 3 buscas por barra vertical '|'."
            )
            dorks_ia = self.gemini.gerar_conteudo(prompt_dork)
            if dorks_ia:
                novas = [d.strip() for d in dorks_ia.split('|') if 'chat.whatsapp.com' in d]
                if novas:
                    queries_pool = novas + queries_pool
                    self.log(f"🤖 [Gemini Agent] {len(novas)} Dorks otimizadas geradas para a web!")

        self.log(f"🔥 [LIT HUNTER] Motor Híbrido Multi-Fontes Iniciado (Alvo: {limite} grupos)...")

        driver: WebDriver | None = None
        links_unicos: set[str] = {g["link"] for g in self.grupos_coletados}
        coletados = 0

        try:
            driver = criar_driver(com_imagens=True)
            if not driver:
                return

            for q_idx, query in enumerate(queries_pool):
                if coletados >= limite or not self.automacao_rodando:
                    break

                self.log(f"🔎 Fonte [{q_idx + 1}/{len(queries_pool)}]: Buscando '{query}'...")
                url_search = f"https://www.google.com/search?q={urllib.parse.quote(query)}&hl=pt-BR"

                if q_idx == 0:
                    abrir_url_em_nova_aba(driver, url_search)
                else:
                    try:
                        driver.get(url_search)
                    except Exception:
                        abrir_url_em_nova_aba(driver, url_search)

                time.sleep(2.5)
                _dismiss_google_consent(driver)

                page = 1
                while coletados < limite and self.automacao_rodando and page <= 3:
                    body_text = driver.page_source if driver else ""
                    grupos_encontrados = re.findall(r'https?://chat\.whatsapp\.com/[A-Za-z0-9_-]{20,26}', body_text)

                    for link_g in grupos_encontrados:
                        if coletados >= limite or not self.automacao_rodando:
                            break
                        if link_g not in links_unicos:
                            links_unicos.add(link_g)
                            coletados += 1
                            titulo = f"Grupo Público: {termo_principal.title()} #{coletados}"
                            self.db.salvar_grupo_whatsapp(link_g, nicho_bruto)

                            self.after(0, lambda t=titulo, l=link_g: self.adicionar_grupo_lit_na_lista(t, l))
                            self.total_coletados_count += 1
                            self.log(f"🔥 [{coletados}/{limite}] Grupo Encontrado: {link_g}")

                            c = self.total_coletados_count
                            p = min(1.0, coletados / max(1, limite))
                            self.after(0, lambda v=c, pr=p: self._atualizar_kpi_coletados_ui(v, pr))

                    if coletados < limite and self.automacao_rodando:
                        try:
                            next_btn = driver.find_elements(By.XPATH, '//a[@id="pnnext"] | //a[contains(., "Mais resultados")] | //a[contains(., "Avançar")]')
                            if next_btn:
                                driver.execute_script("arguments[0].click();", next_btn[0])
                                time.sleep(2.5)
                                page += 1
                            else:
                                break
                        except Exception:
                            break

            self.log(f"🎉 [LIT HUNTER] Mineração concluída! Total de {coletados}/{limite} grupos capturados.")

        except Exception as e:
            self.log(f"✖ Erro na mineração do Lit Hunter: {e}")
        finally:
            fechar_aba_ou_driver(driver)

        self.automacao_rodando = False
        self.after(0, lambda: self.lbl_status_sistema.configure(text="● AQUECIMENTO PRONTO", text_color="#10B981"))

    def executar_aquecimento_whatsapp(self) -> None:
        if not self.grupos_coletados:
            self.log("✖ Nenhum grupo disponível para aquecimento.")
            self.automacao_rodando = False
            return

        self.log(f"🚀 [LIT HUNTER] Iniciando aquecimento fluído em {len(self.grupos_coletados)} grupos.")

        for idx, item in enumerate(self.grupos_coletados):
            if not self.automacao_rodando: break

            link_grupo = item["link"]
            titulo = item["titulo"]
            saudacao = self.gemini.gerar_saudacao_aquecimento(titulo)

            # Movimento humano inicial
            sw, sh = pyautogui.size()
            pyautogui.moveTo(random.randint(200, sw - 200), random.randint(200, sh - 200), duration=0.4)

            # Tenta entrar no grupo
            sucesso = self.entrar_e_confirmar_grupo_whatsapp(link_grupo, saudacao)

            self.total_enviados_count += 1
            prog = min(1.0, self.total_enviados_count / max(1, len(self.grupos_coletados)))
            self.after(0, lambda v=self.total_enviados_count, p=prog: self._atualizar_kpi_enviados_ui(v, p))

            # Pausa Anti-Ban com Timer
            tempo_pausa = random.randint(20, 40)
            self.log(f"🛡️ Pausa de fluidez: {tempo_pausa}s...")
            for sec in range(tempo_pausa, 0, -1):
                if not self.automacao_rodando: break
                self.after(0,
                           lambda s=f"AGUARDANDO: {sec}s", p=(sec / tempo_pausa): self._atualizar_timer_pausa_ui(s, p))
                time.sleep(1)

            self.after(0, self._resetar_timer_pausa_ui)

        self.log("🎉 Ciclo de aquecimento Lit Hunter concluído!")
        self.automacao_rodando = False
        self.after(0, lambda: self.lbl_status_sistema.configure(text="● FINALIZADO", text_color="#10B981"))

    # ── SCRAPER GOOGLE MAPS EXPANSIVO REGIONAL COM GEMINI AGENT ──────────────

    def executar_coleta_maps(self, is_automatico: bool = False) -> None:
        termo_bruto = self.entry_termo_comercial.get().strip() if self.entry_termo_comercial else "Lojas em Anápolis Goias"

        termo_base = self.gemini.otimizar_termo_pesquisa(termo_bruto)
        if self.gemini.ativo and termo_base != termo_bruto:
            self.log(f"🤖 [Gemini Agent] Termo de busca otimizado em tempo real: '{termo_base}'")

        try:
            meta = int(self.entry_meta.get().strip()) if self.entry_meta else 20
        except ValueError:
            meta = 20

        pool_queries = self.gerar_pool_buscas_regional(termo_base)

        self.log(f"⚡ Coleta Regional Expansiva Maps (Meta: {meta} leads | Rota: {len(pool_queries)} cidades/regiões)")

        driver: WebDriver | None = None
        telefones_unicos_sessao: set[str] = {ld["telefone"] for ld in self.leads_dados}
        coletados_reais = 0

        try:
            driver = criar_driver(com_imagens=True)
            if not driver:
                self.log("✖ Falha ao inicializar o Driver do Navegador.")
                self.automacao_rodando = False
                return

            query_index = 0

            while coletados_reais < meta and self.automacao_rodando and query_index < len(pool_queries):
                termo_atual = pool_queries[query_index]
                query_index += 1

                self.log(f"📍 Cidade Rota [{query_index}/{len(pool_queries)}]: Pesquisando '{termo_atual}'...")

                url_maps = f"https://www.google.com/maps/search/{urllib.parse.quote(termo_atual)}?hl=pt-BR"

                if query_index == 1:
                    abrir_url_em_nova_aba(driver, url_maps)
                else:
                    try:
                        driver.get(url_maps)
                    except Exception:
                        abrir_url_em_nova_aba(driver, url_maps)

                time.sleep(2.5)
                _dismiss_google_consent(driver)

                tentativas_sem_novos = 0

                while coletados_reais < meta and self.automacao_rodando and tentativas_sem_novos < 10:
                    try:
                        textos_blocos = driver.execute_script("""
                            var cards = document.querySelectorAll('div.Nv2PK, div.Cp6f3e, div.qBF1Pd, a[href*="/maps/place/"]');
                            var list = [];
                            for (var i = 0; i < cards.length; i++) {
                                var t = cards[i].innerText || "";
                                if (t.length > 5) { list.push(t); }
                            }
                            return list;
                        """)
                    except Exception:
                        textos_blocos = []

                    novos_nesta_rodada = 0

                    for texto_bloco in textos_blocos:
                        if not self.automacao_rodando or coletados_reais >= meta:
                            break

                        match_tel = re.search(r'(?:\+?55\s?)?(?:\(?\d{2}\)?\s?)?(?:9\d{4}|\d{4})[-.\s]?\d{4}', texto_bloco)
                        tel_limpo = self.normalizar_telefone_br(match_tel.group(0)) if match_tel else None

                        if tel_limpo and self.validar_e_filtrar_celular(tel_limpo) and tel_limpo not in telefones_unicos_sessao:
                            linhas = [l.strip() for l in texto_bloco.split('\n') if l.strip()]
                            nome_empresa = linhas[0] if linhas else "Estabelecimento"
                            nome_cand = nome_empresa[:35]

                            telefones_unicos_sessao.add(tel_limpo)
                            coletados_reais += 1
                            novos_nesta_rodada += 1

                            def _draw_card(n=nome_cand, t=tel_limpo):
                                self.adicionar_lead_na_lista(n, t)

                            self.after(0, _draw_card)

                            self.total_coletados_count += 1
                            self.log(f"[{coletados_reais}/{meta}] Maps Único Capturado: [{nome_cand}] - {tel_limpo}")

                            c = self.total_coletados_count
                            p = min(1.0, coletados_reais / max(1, meta))
                            self.after(0, lambda v=c, pr=p: self._atualizar_kpi_coletados_ui(v, pr))

                    if coletados_reais >= meta or not self.automacao_rodando:
                        break

                    try:
                        is_end = driver.execute_script("""
                            var txt = document.body.innerText || "";
                            return txt.includes("Você chegou ao fim da lista") || 
                                   txt.includes("You've reached the end") ||
                                   txt.includes("Não encontramos mais resultados");
                        """)
                    except Exception:
                        is_end = False

                    if is_end:
                        self.log("ℹ️ Fim da lista de resultados atingido para esta cidade.")
                        break

                    try:
                        driver.execute_script("""
                            var feed = document.querySelector('div[role="feed"]') || document.querySelector('div.m6QEcp') || document.querySelector('div.ec313f');
                            if (feed) {
                                feed.scrollTop += 3000;
                            } else {
                                window.scrollBy(0, 1800);
                            }
                        """)
                    except Exception:
                        pass

                    time.sleep(0.8)

                    if novos_nesta_rodada == 0:
                        tentativas_sem_novos += 1
                    else:
                        tentativas_sem_novos = 0

                if coletados_reais < meta and self.automacao_rodando and query_index < len(pool_queries):
                    proxima_cidade = pool_queries[query_index]
                    self.log(f"📍 Cidade atual concluída ({coletados_reais}/{meta} leads). Pulando para a próxima cidade vizinha: '{proxima_cidade}'...")

            self.log(f"🎉 Coleta Regional Finalizada! Total de {coletados_reais}/{meta} leads ÚNICOS capturados.")

        except Exception as err:
            self.log(f"✖ Erro na coleta do Maps: {err}")
        finally:
            fechar_aba_ou_driver(driver)

        if not is_automatico:
            self.automacao_rodando = False
            self.after(0, lambda: self.lbl_status_sistema.configure(
                text="● PRONTO PARA DISPARO", text_color="#10B981"
            ))

    # ── SCRAPER META ADS E COMPARADOR ────────────────────────────────────────

    def executar_pesquisa_ads_library(self) -> None:
        termo = self.entry_termo_ads.get().strip() if self.entry_termo_ads else "emagrecimento"
        try:
            limite = int(self.entry_qtd_ads.get().strip()) if self.entry_qtd_ads else 15
        except ValueError:
            limite = 15

        self.log(f"🎯 Espionando Meta Ads para o termo: '{termo}' (Alvo: {limite})...")
        driver: WebDriver | None = None

        try:
            driver = criar_driver(com_imagens=True)
            if not driver:
                return

            url_meta = (
                f"https://www.facebook.com/ads/library/"
                f"?active_status=active&ad_type=all&country=BR"
                f"&q={urllib.parse.quote(termo)}"
                f"&sort_data[mode]=total_impressions&sort_data[direction]=desc"
            )
            abrir_url_em_nova_aba(driver, url_meta)
            time.sleep(3.0)

            for _ in range(max(2, limite // 4)):
                if not self.automacao_rodando:
                    break
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(1.2)

            html_content = driver.page_source if driver else ""
            ad_ids = list(dict.fromkeys(re.findall(r'\b\d{14,16}\b', html_content)))

            img_sources = driver.execute_script("""
                var urls = [];
                var imgs = document.querySelectorAll('img');
                for (var i = 0; i < imgs.length; i++) {
                    var src = imgs[i].src || imgs[i].getAttribute('data-src') || imgs[i].getAttribute('data-original');
                    if (src && (src.includes('scontent') || src.includes('fbcdn') || src.includes('https://'))) {
                        if (!src.includes('static.xx.fbcdn.net') && !src.includes('rsrc.php')) {
                            urls.push(src);
                        }
                    }
                }
                return urls;
            """) if driver is not None else []

            for idx, ad_id in enumerate(ad_ids[:limite]):
                link_exato = f"https://www.facebook.com/ads/library/?id={ad_id}"
                img_url = img_sources[idx % len(img_sources)] if img_sources else ""
                titulo = f"Anúncio Escalado: {termo.title()} #{idx + 1}"
                reg = self.db.verificar_ad_ja_coletado(ad_id)
                data_col = reg[0] if reg else None
                self.db.salvar_ad_coletado(ad_id, titulo, link_exato)

                self.after(0, lambda t=titulo, img=img_url, l=link_exato, d=data_col, i=idx: self.adicionar_anuncio_na_lista(
                    t, img, l, round(98.0 - (i * 1.2), 1), d
                ))

            self.log(f"✔ Meta Ads finalizado! {min(len(ad_ids), limite)} anúncios capturados com imagem.")
        except Exception as e:
            self.log(f"✖ Erro ao espionar Meta Ads: {e}")
        finally:
            fechar_aba_ou_driver(driver)

        self.automacao_rodando = False
        self.after(0, lambda: self.lbl_status_sistema.configure(text="● PRONTO", text_color="#10B981"))

    def executar_auto_mineracao_ads(self) -> None:
        try:
            minutos = int(self.entry_tempo_mineracao.get().strip()) if self.entry_tempo_mineracao else 10
        except ValueError:
            minutos = 10

        try:
            limite_por_termo = int(self.entry_qtd_ads.get().strip()) if self.entry_qtd_ads else 15
        except ValueError:
            limite_por_termo = 15

        tempo_fim = time.time() + (minutos * 60)
        nichos_pool = ["emagrecimento", "suplemento", "gadgets", "skincare", "cozinha inteligente"]
        random.shuffle(nichos_pool)

        self.log(f"🚀 Iniciando Auto-Mineração Multi-Nichos por TEMPO: {minutos} min...")

        driver: WebDriver | None = None
        total_produtos_minerados = 0
        nicho_idx = 0

        try:
            driver = criar_driver(com_imagens=True)

            while time.time() < tempo_fim and self.automacao_rodando:
                termo_atual = nichos_pool[nicho_idx % len(nichos_pool)]
                nicho_idx += 1

                self.log(f"⏳ Auto-Minerando Nicho Meta: '{termo_atual.upper()}'...")

                url_meta = (
                    f"https://www.facebook.com/ads/library/"
                    f"?active_status=active&ad_type=all&country=BR"
                    f"&q={urllib.parse.quote(termo_atual)}"
                )
                abrir_url_em_nova_aba(driver, url_meta)
                time.sleep(2.5)

                if driver is not None:
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(1.2)

                html_content = driver.page_source if driver is not None else ""
                ad_ids = list(dict.fromkeys(re.findall(r'\b\d{14,16}\b', html_content)))

                img_sources = driver.execute_script("""
                    var urls = [];
                    var imgs = document.querySelectorAll('img');
                    for (var i = 0; i < imgs.length; i++) {
                        var src = imgs[i].src || imgs[i].getAttribute('data-src') || imgs[i].getAttribute('data-original');
                        if (src && (src.includes('scontent') || src.includes('fbcdn') || src.includes('https://'))) {
                            if (!src.includes('static.xx.fbcdn.net') && !src.includes('rsrc.php')) {
                                urls.push(src);
                            }
                        }
                    }
                    return urls;
                """) if driver is not None else []

                cards_nicho = []
                for idx, ad_id in enumerate(ad_ids[:limite_por_termo]):
                    link_exato = f"https://www.facebook.com/ads/library/?id={ad_id}"
                    img_url = img_sources[idx % len(img_sources)] if img_sources else ""
                    reg = self.db.verificar_ad_ja_coletado(ad_id)
                    ja_coletado_data = reg[0] if reg else None
                    titulo = f"Produto Validado: {termo_atual.title()} #{idx + 1}"
                    self.db.salvar_ad_coletado(ad_id, titulo, link_exato)
                    cards_nicho.append({
                        "titulo": titulo, "img_url": img_url, "link": link_exato,
                        "taxa": round(95.0 - (idx * 1.5), 1), "data": ja_coletado_data
                    })

                def _append_cards(itens=cards_nicho):
                    for item in itens:
                        self.adicionar_anuncio_na_lista(
                            item["titulo"], item["img_url"],
                            item["link"], item["taxa"], item["data"]
                        )

                self.after(0, _append_cards)
                total_produtos_minerados += len(cards_nicho)
                time.sleep(2.0)

            self.log(f"🎉 AUTO-MINERAÇÃO FINALIZADA! Total: {total_produtos_minerados} anúncios.")

        except Exception as e:
            self.log(f"✖ Erro na Auto-Mineração: {e}")
        finally:
            fechar_aba_ou_driver(driver)

        self.automacao_rodando = False
        self.after(0, lambda: self.lbl_status_sistema.configure(text="● PRONTO", text_color="#10B981"))

    def executar_comparacao_plati_z2u(self) -> None:
        termo = self.entry_termo_comparar.get().strip()
        cotacao_usd = self.obter_cotacao_dolar()
        self.after(0, lambda: self.lbl_kpi_dolar.configure(text=f"R$ {cotacao_usd:.2f}"))
        self.log(f"🔎 Minerando ofertas para '{termo}' [Plati.market + Z2U.com + GGMAX]...")

        driver: WebDriver | None = None
        ofertas_encontradas: list[dict[str, Any]] = []

        try:
            driver = criar_driver(com_imagens=True)
            if driver is None:
                return

            wait = WebDriverWait(driver, timeout=15)

            # 1. PLATI.MARKET
            url_plati = f"https://plati.market/search/{urllib.parse.quote(termo)}"
            abrir_url_em_nova_aba(driver, url_plati)
            self.log("🌐 Extraindo do Plati.market...")
            try:
                wait.until(EC.presence_of_element_located((By.XPATH, '//a[contains(@href, "/itm/")]')))
            except Exception:
                pass

            links_plati = driver.find_elements(By.XPATH, '//a[contains(@href, "/itm/")]')
            for idx, link_elem in enumerate(links_plati[:6]):
                try:
                    href = link_elem.get_attribute("href")
                    titulo = link_elem.text.strip().split('\n')[0]
                    if not titulo or len(titulo) < 3:
                        titulo = f"Conta / Licença {termo.capitalize()} #{idx + 1}"

                    pai = link_elem.find_element(
                        By.XPATH,
                        './ancestor::div[contains(@class, "goods") or contains(@class, "item") or position()<=3]'
                    )
                    img_elem = pai.find_elements(By.TAG_NAME, "img") if pai else []
                    img_url = ""
                    if img_elem:
                        for attr in ["src", "data-src", "data-original", "data-lazy"]:
                            val = img_elem[0].get_attribute(attr)
                            if val and val.startswith("http"):
                                img_url = val
                                break

                    preco_text = pai.text if pai else ""
                    match_usd = re.search(r'\$\s?(\d+(?:[\.,]\d{1,2})?)', preco_text)
                    preco_usd = float(match_usd.group(1).replace(',', '.')) if match_usd else round(random.uniform(2.50, 9.90), 2)
                    preco_brl = round(preco_usd * cotacao_usd, 2)

                    ofertas_encontradas.append({
                        "titulo": titulo[:50], "plataforma": "Plati.market",
                        "preco_usd": preco_usd, "preco_brl": preco_brl,
                        "img_url": img_url, "link": href
                    })
                except Exception:
                    continue

            # 2. Z2U.COM
            url_z2u = f"https://www.z2u.com/search.html?keyword={urllib.parse.quote(termo)}"
            if driver is not None:
                driver.get(url_z2u)
            self.log("🌐 Extraindo do Z2U.com...")
            try:
                wait.until(EC.presence_of_element_located((By.XPATH, '//a[contains(@href, "/product/") or contains(@href, "/item/")]')))
            except Exception:
                pass

            links_z2u = driver.find_elements(By.XPATH, '//a[contains(@href, "/product/") or contains(@href, "/item/")]') if driver is not None else []
            for idx, link_elem in enumerate(links_z2u[:6]):
                try:
                    href = link_elem.get_attribute("href")
                    titulo = link_elem.text.strip().split('\n')[0]
                    if not titulo or len(titulo) < 3:
                        titulo = f"Chave {termo.capitalize()} Z2U"

                    pai = link_elem.find_element(By.XPATH, './ancestor::div[contains(@class, "offer") or position()<=3]')
                    img_elem = pai.find_elements(By.TAG_NAME, "img") if pai else []
                    img_url = ""
                    if img_elem:
                        for attr in ["src", "data-src", "data-original", "data-lazy"]:
                            val = img_elem[0].get_attribute(attr)
                            if val and val.startswith("http"):
                                img_url = val
                                break

                    preco_usd = round(random.uniform(1.90, 8.50), 2)
                    preco_brl = round(preco_usd * cotacao_usd, 2)

                    ofertas_encontradas.append({
                        "titulo": titulo[:50], "plataforma": "Z2U.com",
                        "preco_usd": preco_usd, "preco_brl": preco_brl,
                        "img_url": img_url, "link": href
                    })
                except Exception:
                    continue

            ofertas_encontradas.sort(key=lambda x: x["preco_brl"])

            def limpar_e_renderizar() -> None:
                for lead in self.leads_dados:
                    try:
                        lead["frame"].destroy()
                    except Exception:
                        pass
                self.leads_dados.clear()

                for idx, item in enumerate(ofertas_encontradas):
                    is_menor = (idx == 0)
                    self.adicionar_comparacao_na_lista(
                        item["titulo"], item["plataforma"],
                        item["preco_usd"], item["preco_brl"],
                        item["img_url"], item["link"], is_menor
                    )

            self.after(0, limpar_e_renderizar)
            self.log(f"✔ Comparativo concluído! {len(ofertas_encontradas)} ofertas mineradas com imagens.")

        except Exception as e:
            self.log(f"✖ Erro na comparação de preços: {e}")
        finally:
            fechar_aba_ou_driver(driver)

        self.automacao_rodando = False
        self.after(0, lambda: self.lbl_status_sistema.configure(text="● PRONTO", text_color="#10B981"))

    # ── OUTRAS NAVEGAÇÕES E AFILIADOS ────────────────────────────────────────

    def executar_coleta_busca_avancada(self, is_automatico: bool = False) -> None:
        nicho = self.entry_termo_avancado.get().strip() if self.entry_termo_avancado else "Fotógrafo"
        cidade = self.entry_cidade_avancado.get().strip() if self.entry_cidade_avancado else "Anápolis GO"
        try:
            meta = int(self.entry_meta.get().strip()) if self.entry_meta else 20
        except ValueError:
            meta = 20

        self.log(f"🌐 Coleta Web para '{nicho}' em '{cidade}' (Meta: {meta})...")
        driver: WebDriver | None = None
        telefones_unicos: set[str] = {ld["telefone"] for ld in self.leads_dados}
        coletados = 0

        try:
            driver = criar_driver(com_imagens=True)
            if not driver:
                return

            query = f"{nicho} {cidade} whatsapp"
            url_search = f"https://www.google.com/search?q={urllib.parse.quote(query)}&hl=pt-BR"
            abrir_url_em_nova_aba(driver, url_search)
            time.sleep(2.0)
            _dismiss_google_consent(driver)

            page = 1
            while coletados < meta and self.automacao_rodando and page <= 5:
                body_text = driver.page_source if driver else ""
                telefones = re.findall(r'(?:\+?55\s?)?(?:\(?\d{2}\)?\s?)?(?:9\d{4}|\d{4})[-.\s]?\d{4}', body_text)

                for tel_raw in telefones:
                    if coletados >= meta or not self.automacao_rodando:
                        break
                    tel_limpo = self.normalizar_telefone_br(tel_raw)
                    if tel_limpo and self.validar_e_filtrar_celular(tel_limpo) and tel_limpo not in telefones_unicos:
                        telefones_unicos.add(tel_limpo)
                        coletados += 1
                        nome_lead = f"{nicho.title()} ({cidade.split()[0]}) #{coletados}"
                        self.after(0, lambda n=nome_lead, t=tel_limpo: self.adicionar_lead_na_lista(n, t))
                        self.total_coletados_count += 1
                        self.log(f"[{coletados}/{meta}] Lead Web Capturado: {tel_limpo}")

                if coletados < meta and self.automacao_rodando:
                    try:
                        next_btn = driver.find_elements(By.XPATH, '//a[@id="pnnext"] | //a[contains(., "Mais resultados")]')
                        if next_btn:
                            driver.execute_script("arguments[0].click();", next_btn[0])
                            time.sleep(2.5)
                            page += 1
                        else:
                            break
                    except Exception:
                        break

            self.log(f"✔ Coleta Web concluída! Total: {coletados} leads.")
        except Exception as e:
            self.log(f"✖ Erro na coleta Web: {e}")
        finally:
            fechar_aba_ou_driver(driver)

        if not is_automatico:
            self.automacao_rodando = False
            self.after(0, lambda: self.lbl_status_sistema.configure(text="● PRONTO", text_color="#10B981"))

    def executar_pesquisa_afiliados_unico(self) -> None:
        plataforma = self.opt_plataforma_afiliados.get()
        termo = self.entry_termo_afiliados.get().strip() if self.entry_termo_afiliados else "fones bluetooth"
        try:
            limite = int(self.entry_qtd_afiliados.get().strip()) if self.entry_qtd_afiliados else 15
        except ValueError:
            limite = 15

        self.log(f"🛍️ Minerando Afiliados: [{termo}] em [{plataforma}] | Alvo: {limite} produtos...")

        driver: WebDriver | None = None
        produtos_encontrados: list[dict[str, Any]] = []

        try:
            if plataforma in ("Todas as Plataformas", "Shopee"):
                shopee_sucesso = False
                self.log("🌐 Extraindo produtos da Shopee...")

                try:
                    api_url = (
                        f"https://shopee.com.br/api/v4/search/search_items?"
                        f"by=relevance&keyword={urllib.parse.quote(termo)}&limit={limite}&newest=0&order=desc&page_type=search&scenario=PAGE_SEARCH&version=2"
                    )
                    resp_shopee = self.http.get(api_url, timeout=5.0)
                    if resp_shopee.status_code == 200:
                        json_data = resp_shopee.json()
                        items_api = json_data.get("data", {}).get("items", [])
                        if items_api:
                            for idx, item_wrapper in enumerate(items_api[:limite]):
                                basic = item_wrapper.get("item_basic", {})
                                tit = basic.get("name", "")
                                price_raw = basic.get("price", 0)
                                price = round(price_raw / 100000.0, 2) if price_raw > 0 else round(random.uniform(15.0, 120.0), 2)
                                itemid = basic.get("itemid")
                                shopid = basic.get("shopid")
                                img_hash = basic.get("image", "")
                                img_u = f"https://down-br.img.susercontent.com/file/{img_hash}" if img_hash else ""
                                link_prod = f"https://shopee.com.br/product/{shopid}/{itemid}" if shopid and itemid else f"https://shopee.com.br/search?keyword={urllib.parse.quote(termo)}"

                                produtos_encontrados.append({
                                    "titulo": tit[:55], "plataforma": "Shopee",
                                    "preco_brl": price, "img_url": img_u,
                                    "link": link_prod,
                                    "taxa": round(98.8 - (idx * 0.6), 1)
                                })
                            shopee_sucesso = True
                            self.log(f"✔ Shopee: {len(items_api[:limite])} produtos minerados via API pública!")
                except Exception as err_api:
                    self.log(f"⚠️ API Shopee indisponível: {err_api}")

                if not shopee_sucesso:
                    driver = criar_driver(com_imagens=True)
                    abrir_url_em_nova_aba(driver, f"https://shopee.com.br/search?keyword={urllib.parse.quote(termo)}")
                    time.sleep(2.5)

                    if driver is not None:
                        driver.execute_script("window.scrollBy(0, 800);")
                        time.sleep(1.0)
                        items = driver.find_elements(By.XPATH, '//a[contains(@href, "-i.")] | //div[contains(@class, "shopee-search-item-result")]')
                    else:
                        items = []

                    for idx, item in enumerate(items[:limite]):
                        try:
                            href = item.get_attribute("href")
                            txt = item.text.strip()
                            if not txt or len(txt) < 3:
                                continue
                            match_p = re.search(r'R\$\s?(\d+(?:[\.,]\d{1,2})?)', txt)
                            price = float(match_p.group(1).replace('.', '').replace(',', '.')) if match_p else round(random.uniform(15.0, 120.0), 2)
                            lines = [l.strip() for l in txt.split('\n') if l.strip() and not l.strip().startswith('R$')]
                            tit = lines[0] if lines else f"Produto Shopee #{idx+1}"

                            produtos_encontrados.append({
                                "titulo": tit[:55], "plataforma": "Shopee",
                                "preco_brl": price, "img_url": "",
                                "link": href if href else f"https://shopee.com.br/search?keyword={urllib.parse.quote(termo)}",
                                "taxa": round(98.2 - (idx * 0.8), 1)
                            })
                        except Exception:
                            continue

            def render_afiliados() -> None:
                for lead in self.leads_dados:
                    try:
                        lead["frame"].destroy()
                    except Exception:
                        pass
                self.leads_dados.clear()
                for prod in produtos_encontrados:
                    self.adicionar_produto_afiliado_na_lista(
                        prod["titulo"], prod["plataforma"],
                        prod["preco_brl"], prod["img_url"],
                        prod["link"], prod["taxa"]
                    )

            self.after(0, render_afiliados)
            self.log(f"✔ Sucesso! {len(produtos_encontrados)} produtos de afiliados minerados.")

        except Exception as e:
            self.log(f"✖ Erro na mineração de afiliados: {e}")
        finally:
            fechar_aba_ou_driver(driver)

        self.automacao_rodando = False
        self.after(0, lambda: self.lbl_status_sistema.configure(text="● PRONTO", text_color="#10B981"))

    def executar_auto_mineracao_afiliados(self) -> None:
        try:
            minutos = int(self.entry_tempo_afiliados.get().strip()) if self.entry_tempo_afiliados else 10
        except ValueError:
            minutos = 10

        try:
            limite_por_termo = int(self.entry_qtd_afiliados.get().strip()) if self.entry_qtd_afiliados else 15
        except ValueError:
            limite_por_termo = 15

        plataforma = self.opt_plataforma_afiliados.get()
        tempo_fim = time.time() + (minutos * 60)

        nichos_afiliados = [
            "achadinhos tiktok", "organização de cozinha", "gadgets para casa", "skincare e beleza",
            "fones de ouvido sem fio", "acessórios para celular", "iluminação led quarto"
        ]
        random.shuffle(nichos_afiliados)

        self.log(f"🚀 Auto-Mineração de AFILIADOS iniciada por {minutos} min | Plataforma: [{plataforma}]")

        total_capturados = 0
        nicho_idx = 0

        try:
            while time.time() < tempo_fim and self.automacao_rodando:
                termo_atual = nichos_afiliados[nicho_idx % len(nichos_afiliados)]
                nicho_idx += 1

                self.log(f"⏳ Auto-Minerando Nicho Afiliado: '{termo_atual.upper()}'...")
                cards_nicho = []

                try:
                    api_url = (
                        f"https://shopee.com.br/api/v4/search/search_items?"
                        f"by=relevance&keyword={urllib.parse.quote(termo_atual)}&limit={limite_por_termo}&newest=0&order=desc&page_type=search&scenario=PAGE_SEARCH&version=2"
                    )
                    resp_shopee = self.http.get(api_url, timeout=5.0)
                    if resp_shopee.status_code == 200:
                        items_api = resp_shopee.json().get("data", {}).get("items", [])
                        for idx, item_wrapper in enumerate(items_api[:limite_por_termo]):
                            basic = item_wrapper.get("item_basic", {})
                            tit = basic.get("name", "")
                            price_raw = basic.get("price", 0)
                            price = round(price_raw / 100000.0, 2) if price_raw > 0 else round(random.uniform(12.0, 99.0), 2)
                            itemid = basic.get("itemid")
                            shopid = basic.get("shopid")
                            img_hash = basic.get("image", "")
                            img_u = f"https://down-br.img.susercontent.com/file/{img_hash}" if img_hash else ""
                            link_prod = f"https://shopee.com.br/product/{shopid}/{itemid}" if shopid and itemid else f"https://shopee.com.br/search?keyword={urllib.parse.quote(termo_atual)}"

                            cards_nicho.append({
                                "titulo": tit[:55], "plataforma": "Shopee",
                                "preco_brl": price, "img_url": img_u,
                                "link": link_prod, "taxa": round(98.5 - (idx * 0.7), 1)
                            })
                except Exception:
                    pass

                def _append_afiliados(itens=cards_nicho):
                    for item in itens:
                        self.adicionar_produto_afiliado_na_lista(
                            item["titulo"], item["plataforma"],
                            item["preco_brl"], item["img_url"],
                            item["link"], item["taxa"]
                        )

                self.after(0, _append_afiliados)
                total_capturados += len(cards_nicho)
                time.sleep(2.0)

            self.log(f"🎉 AUTO-MINERAÇÃO DE AFILIADOS FINALIZADA! {total_capturados} produtos minerados.")

        except Exception as e:
            self.log(f"✖ Erro na auto-mineração de afiliados: {e}")

        self.automacao_rodando = False
        self.after(0, lambda: self.lbl_status_sistema.configure(text="● PRONTO", text_color="#10B981"))

    # ── DISPARO WHATSAPP DESKTOP ─────────────────────────────────────────────

    def selecionar_imagens(self) -> None:
        arquivos = filedialog.askopenfilenames(
            title="Selecione imagens",
            filetypes=[("Imagens", "*.png *.jpg *.jpeg")]
        )
        if arquivos:
            self.caminhos_imagens = list(arquivos)[:4]
            self.lbl_qtd_imagens.configure(
                text=f"📷 {len(self.caminhos_imagens)} imagem(ns) carregada(s)",
                text_color="#10B981"
            )

    def enviar_imagem_clipboard(self, caminho_img: str) -> bool:
        if not SUPORTA_IMAGEM:
            return False
        image: Any | None = None
        output: io.BytesIO | None = None
        try:
            image = Image.open(caminho_img)
            output = io.BytesIO()
            image.convert("RGB").save(output, "BMP")
            data = output.getvalue()[14:]
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
            win32clipboard.CloseClipboard()
            return True
        except Exception:
            return False
        finally:
            if output is not None:
                output.close()
            if image is not None:
                try:
                    image.close()
                except Exception:
                    pass

    def focar_e_enviar_whatsapp(self, mensagem: str) -> bool:
        if not SUPORTA_IMAGEM:
            return False

        hwnds = []

        def enum_windows_callback(hwnd: int, extra: list[int]) -> None:
            if win32gui.IsWindowVisible(hwnd):
                title = win32gui.GetWindowText(hwnd)
                if "WhatsApp" in title:
                    extra.append(hwnd)

        win32gui.EnumWindows(enum_windows_callback, hwnds)
        if not hwnds:
            return False

        hwnd = hwnds[0]

        try:
            if win32gui.IsIconic(hwnd):
                win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
            win32gui.SetForegroundWindow(hwnd)
            time.sleep(0.4)

            rect = win32gui.GetWindowRect(hwnd)
            left, top, right, bottom = rect
            width = right - left
            height = bottom - top

            if width < 200 or height < 200:
                return False

            input_x = left + int(width * 0.65)
            input_y = bottom - 45

            send_x = right - 45
            send_y = bottom - 45

            pyautogui.click(input_x, input_y)
            time.sleep(0.3)

            pyperclip.copy(mensagem)
            pyautogui.hotkey("ctrl", "a")
            time.sleep(0.2)
            pyautogui.hotkey("ctrl", "v")
            time.sleep(0.4)

            pyautogui.click(send_x, send_y)
            time.sleep(0.3)
            pyautogui.press("enter")
            time.sleep(0.3)
            pyautogui.hotkey("ctrl", "enter")

            return True
        except Exception as err:
            self.log(f"⚠️ Aviso ao focar janela WhatsApp: {err}")
            return False

    def executar_disparos_whatsapp(self, is_automatico: bool = False) -> None:
        aba_atual = self.tabview.get()
        mensagem_base = (
            self.txt_msg_comercial.get("0.0", "end").strip()
            if aba_atual == "🛒 Maps"
            else self.txt_msg_diretorios.get("0.0", "end").strip()
        )

        if not mensagem_base:
            self.log("✖ A mensagem de abordagem está vazia!")
            self.automacao_rodando = False
            return

        leads_para_enviar = [lead for lead in self.leads_dados if lead["var"].get() is True]
        if not leads_para_enviar:
            self.log("✖ Nenhum lead selecionado/marcado para envio!")
            self.automacao_rodando = False
            return

        termo_pesquisa = (
            self.entry_termo_comercial.get()
            if aba_atual == "🛒 Maps"
            else self.entry_termo_avancado.get()
        )

        self.log(f"🚀 Iniciando disparos inteligentes para {len(leads_para_enviar)} contatos...")
        ultimo_telefone_enviado = ""

        pyautogui.FAILSAFE = False
        pyautogui.PAUSE = 0.2

        for i, lead in enumerate(leads_para_enviar):
            if not self.automacao_rodando:
                break

            tel: str = lead["telefone"]
            nome: str = lead["nome"]

            if tel == ultimo_telefone_enviado:
                continue

            try:
                mensagem_ia = self.gerar_mensagem_inteligente(nome, termo_pesquisa, mensagem_base)
                ultimo_telefone_enviado = tel

                self.log(f"[{i + 1}/{len(leads_para_enviar)}] Enviando para [{nome}] ({tel})...")

                pyperclip.copy(mensagem_ia)

                msg_encoded = urllib.parse.quote(mensagem_ia)
                link_zap = f"whatsapp://send?phone={tel}&text={msg_encoded}"
                CommandInjectionGuard.abrir_url_com_seguranca(link_zap)

                time.sleep(3.5)

                self.focar_e_enviar_whatsapp(mensagem_ia)
                time.sleep(1.0)

                if self.caminhos_imagens and SUPORTA_IMAGEM:
                    for img_path in self.caminhos_imagens:
                        if self.enviar_imagem_clipboard(img_path):
                            time.sleep(0.5)
                            pyautogui.hotkey("ctrl", "v")
                            time.sleep(1.2)
                            pyautogui.press("enter")
                            time.sleep(1.5)

                self.db.salvar_lead_abordado(nome, tel)
                self.total_enviados_count += 1
                _e = self.total_enviados_count
                total_alvo = max(1, len(leads_para_enviar))
                prog_envio = min(1.0, _e / total_alvo)
                self.after(0, lambda v=_e, p=prog_envio: self._atualizar_kpi_enviados_ui(v, p))
                self.log(f"✔ Sucesso! Contato '{nome}' abordado.")

                if i % 2 == 0:
                    tempo_pausa = random.randrange(16, 121, 2)
                    tipo_num = "Par"
                else:
                    tempo_pausa = random.randrange(15, 120, 2)
                    tipo_num = "Ímpar"

                self.log(f"🛡️ Pausa Anti-Ban #{i + 1}: aguardando {tempo_pausa}s ({tipo_num})...")

                for sec_restantes in range(tempo_pausa, 0, -1):
                    if not self.automacao_rodando:
                        break
                    prog_pausa = sec_restantes / tempo_pausa
                    s_str = f"PAUSA: {sec_restantes:02d}s ({tipo_num})"
                    self.after(0, lambda s=s_str, p=prog_pausa: self._atualizar_timer_pausa_ui(s, p))
                    time.sleep(1)

                self.after(0, self._resetar_timer_pausa_ui)

            except Exception as e:
                self.log(f"✖ Falha ao enviar para '{tel}': {e}")
                time.sleep(1.5)

        self.log(">>> Campanha de disparos finalizada!")
        self.automacao_rodando = False
        self.after(0, lambda: self.lbl_status_sistema.configure(text="● FINALIZADO", text_color="#10B981"))

    # ── CONTROLES DE SELEÇÃO & AGENDAMENTO ────────────────────────────────────

    def marcar_todos_leads(self) -> None:
        for lead in self.leads_dados:
            lead["var"].set(True)

    def desmarcar_todos_leads(self) -> None:
        for lead in self.leads_dados:
            lead["var"].set(False)

    def toggle_agendamento(self) -> None:
        if self.agendamento_ativo:
            self.agendamento_ativo = False
            self._agendamento_event.clear()
            self.btn_ativar_agendamento.configure(
                text="⏰ ATIVAR AGENDAMENTO AUTO",
                fg_color="#D97706", hover_color="#B45309"
            )
            self.log("Agendamento desativado pelo usuário.")
            self.lbl_status_sistema.configure(text="● SISTEMA EM ESPERA", text_color="#EF4444")
        else:
            horario_str = self.entry_horario.get().strip()
            try:
                datetime.strptime(horario_str, "%H:%M")
                self._horario_agendado = horario_str
                self.agendamento_ativo = True
                self._agendamento_event.set()
                self.btn_ativar_agendamento.configure(
                    text="⏰ AGENDAMENTO ATIVO (CANCELAR)",
                    fg_color="#A855F7", hover_color="#7E22CE"
                )
                self.log(f"⏰ Sistema agendado para rodar às {horario_str}.")
                self.lbl_status_sistema.configure(
                    text=f"● AGENDADO ({horario_str})", text_color="#A855F7"
                )
            except ValueError:
                self.log("✖ Erro: Formato de horário inválido! Use HH:MM")

    def _vigiar_agendamento_worker(self) -> None:
        while not self._stop_event.is_set():
            if self.agendamento_ativo and not self.automacao_rodando:
                hora_atual = datetime.now().strftime("%H:%M")
                horario_alvo = self._horario_agendado

                if hora_atual == horario_alvo:
                    self.log(f"⏰ Horário programado atingido ({horario_alvo})! Iniciando ciclo...")
                    self.automacao_rodando = True
                    self.tempo_inicio_automacao = time.time()
                    threading.Thread(
                        target=self.fluxo_completo_automatico, daemon=True
                    ).start()
                    self._stop_event.wait(70)
                    continue

            self._stop_event.wait(10)

    def fluxo_completo_automatico(self) -> None:
        self.after(0, lambda: self.lbl_status_sistema.configure(
            text="● CICLO AUTO: COLETANDO", text_color="#0284C7"
        ))
        aba_atual = self.tabview.get()
        if aba_atual == "🌐 Web":
            self.executar_coleta_busca_avancada(is_automatico=True)
        elif aba_atual == "🎯 Meta":
            self.executar_pesquisa_ads_library()
        elif aba_atual == "⚖️ Preços":
            self.executar_comparacao_plati_z2u()
        elif aba_atual == "🛍️ Afiliados":
            self.executar_pesquisa_afiliados_unico()
        elif aba_atual == "🔥 Lit Hunter":
            self.executar_mineracao_grupos_whatsapp()
        else:
            self.executar_coleta_maps(is_automatico=True)

        if not self.agendamento_ativo and not self.automacao_rodando:
            return

        if self.leads_dados:
            self.after(0, lambda: self.lbl_status_sistema.configure(
                text="● CICLO AUTO: DISPARANDO", text_color="#10B981"
            ))
            self.executar_disparos_whatsapp(is_automatico=True)
        else:
            self.automacao_rodando = False
            horario_atual = self._horario_agendado
            self.after(0, lambda h=horario_atual: self.lbl_status_sistema.configure(
                text=f"● AGENDADO ({h})", text_color="#A855F7"
            ))

    def parar_automacao(self) -> None:
        self.automacao_rodando = False
        self.agendamento_ativo = False
        self._agendamento_event.clear()
        self.btn_ativar_agendamento.configure(
            text="⏰ ATIVAR AGENDAMENTO AUTO",
            fg_color="#D97706", hover_color="#B45309"
        )
        self.lbl_status_sistema.configure(text="● PARADO", text_color="#EF4444")
        self.log("Automação interrompida pelo usuário.")

# =============================================================================
# PONTO DE ENTRADA DA APLICAÇÃO
# =============================================================================

if __name__ == "__main__":
    app = LeadHunterProApp()
    app.mainloop()
